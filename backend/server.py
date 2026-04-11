from fastapi import FastAPI, APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, Response, Request, Body
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from emergentintegrations.llm.chat import LlmChat, UserMessage as LlmUserMessage
import tempfile
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from starlette.responses import JSONResponse
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from jose import JWTError, jwt
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import os
import uuid
import logging
import io
import json
import requests
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# Import seed data
try:
    from seed_data import SEED_DATA
except ImportError:
    SEED_DATA = {}

# MongoDB
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT
SECRET_KEY = os.environ.get('JWT_SECRET_KEY', 'your-secret-key-change-in-production')
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Object Storage
STORAGE_URL = "https://integrations.emergentagent.com/objstore/api/v1/storage"
EMERGENT_KEY = os.environ.get("EMERGENT_LLM_KEY")
APP_NAME = "proptech-turkey"
storage_key = None

def init_storage():
    global storage_key
    if storage_key:
        return storage_key
    resp = requests.post(f"{STORAGE_URL}/init", json={"emergent_key": EMERGENT_KEY}, timeout=30)
    resp.raise_for_status()
    storage_key = resp.json()["storage_key"]
    return storage_key

def put_object(path: str, data: bytes, content_type: str) -> dict:
    key = init_storage()
    resp = requests.put(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key, "Content-Type": content_type},
        data=data, timeout=120
    )
    resp.raise_for_status()
    return resp.json()

def get_object(path: str):
    key = init_storage()
    resp = requests.get(
        f"{STORAGE_URL}/objects/{path}",
        headers={"X-Storage-Key": key}, timeout=60
    )
    resp.raise_for_status()
    return resp.content, resp.headers.get("Content-Type", "application/octet-stream")

MIME_TYPES = {
    "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
    "gif": "image/gif", "webp": "image/webp", "pdf": "application/pdf",
    "json": "application/json", "csv": "text/csv", "txt": "text/plain",
    "doc": "application/msword", "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "xls": "application/vnd.ms-excel", "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "kml": "application/vnd.google-earth.kml+xml", "kmz": "application/vnd.google-earth.kmz",
    "geojson": "application/geo+json",
}

app = FastAPI(title="PropTech Turkey API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ============= HELPERS =============

def hash_password(password: str) -> str:
    return pwd_context.hash(password)

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def extract_centroid_from_kml(fext: str, data: bytes) -> dict | None:
    """Extract lat/lng centroid from KML or KMZ bytes."""
    import re
    import xml.etree.ElementTree as ET
    import zipfile as _zf
    import io as _io2
    try:
        kml_bytes = data
        if fext == "kmz":
            with _zf.ZipFile(_io2.BytesIO(data)) as z2:
                kfiles = [n for n in z2.namelist() if n.lower().endswith('.kml')]
                if not kfiles:
                    return None
                kml_bytes = z2.read(kfiles[0])
        kml_text = kml_bytes.decode('utf-8', errors='ignore')
        coords_raw = ' '.join(
            el.text or '' for el in ET.fromstring(kml_text).iter()
            if el.tag.endswith('coordinates')
        )
        pairs = re.findall(r'([-\d.]+),([-\d.]+)', coords_raw)
        if pairs:
            lons = [float(p[0]) for p in pairs]
            lats = [float(p[1]) for p in pairs]
            return {"lat": round(sum(lats)/len(lats), 6), "lng": round(sum(lons)/len(lons), 6)}
    except Exception:
        pass
    return None

def extract_centroid_from_geojson(data: bytes) -> dict | None:
    """Extract lat/lng centroid from GeoJSON bytes."""
    import json as _json
    all_lons: list[float] = []
    all_lats: list[float] = []

    def collect(coords):
        if not coords:
            return
        if isinstance(coords[0], (int, float)):
            if len(coords) >= 2:
                all_lons.append(float(coords[0]))
                all_lats.append(float(coords[1]))
        else:
            for item in coords:
                collect(item)

    def process_geom(geom):
        if not geom:
            return
        gtype = geom.get('type', '')
        if gtype == 'GeometryCollection':
            for g in geom.get('geometries', []):
                process_geom(g)
        else:
            collect(geom.get('coordinates', []))

    try:
        geojson = _json.loads(data.decode('utf-8'))
        t = geojson.get('type', '')
        if t == 'FeatureCollection':
            for f in geojson.get('features', []):
                process_geom(f.get('geometry'))
        elif t == 'Feature':
            process_geom(geojson.get('geometry'))
        else:
            process_geom(geojson)
        if all_lats:
            return {
                "lat": round(sum(all_lats) / len(all_lats), 6),
                "lng": round(sum(all_lons) / len(all_lons), 6),
            }
    except Exception:
        pass
    return None

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"email": email}, {"_id": 0, "password": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def require_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

async def get_session_user(request: Request) -> Optional[dict]:
    """Get user from session cookie or Authorization header."""
    token = request.cookies.get("session_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        return None
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        return None
    expires_at = session.get("expires_at")
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at and expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at and expires_at < datetime.now(timezone.utc):
        return None
    user = await db.app_users.find_one({"user_id": session["user_id"]}, {"_id": 0, "password": 0})
    return user

async def require_session_user(request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user

# ============= MODELS =============

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserRegister(BaseModel):
    full_name: str
    phone: str = ""
    email: EmailStr
    password: str

class UserUpdateAdmin(BaseModel):
    full_name: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None
    status: Optional[str] = None
    admin_note: Optional[str] = None
    tags: Optional[List[str]] = None

class MembershipUpdate(BaseModel):
    membership_plan: Optional[str] = None
    membership_start_at: Optional[str] = None
    membership_end_at: Optional[str] = None
    membership_active: Optional[bool] = None
    extend_days: Optional[int] = None

class PermissionsUpdate(BaseModel):
    permissions: Dict[str, str]

class NoteUpdate(BaseModel):
    admin_note: str = ""
    tags: List[str] = []

class InvestmentCalculation(BaseModel):
    city: str
    district: str
    neighborhood: str
    ada: str
    parsel: str
    land_size_sqm: float
    emsal: float
    construction_cost_per_sqm: float

class InvestmentResult(BaseModel):
    total_construction_area: float
    estimated_apartments: int
    total_construction_cost: float
    estimated_project_value: float
    potential_profit: float
    roi_percentage: float

# ============= ADMIN AUTH =============

@api_router.post("/auth/login")
async def admin_login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    access_token = create_access_token({"sub": credentials.email})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {"email": user["email"], "full_name": user.get("full_name", ""), "role": user["role"]}
    }

@api_router.get("/auth/me")
async def get_me(request: Request):
    # Try session cookie first (app users)
    session_user = await get_session_user(request)
    if session_user:
        return session_user
    # Try JWT bearer token (admin)
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        try:
            payload = jwt.decode(auth[7:], SECRET_KEY, algorithms=[ALGORITHM])
            email = payload.get("sub")
            if email:
                user = await db.users.find_one({"email": email}, {"_id": 0, "password": 0})
                if user:
                    return user
        except JWTError:
            pass
    raise HTTPException(status_code=401, detail="Not authenticated")

# ============= USER REGISTRATION & LOGIN =============

@api_router.post("/auth/register")
async def register_user(data: UserRegister):
    existing = await db.app_users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Bu e-posta adresi zaten kayıtlı")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    user = {
        "user_id": user_id,
        "full_name": data.full_name,
        "phone": data.phone,
        "email": data.email,
        "password": hash_password(data.password),
        "role": "user",
        "auth_provider": "email",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.app_users.insert_one(user)
    # Create session
    session_token = f"sess_{uuid.uuid4().hex}"
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
        "created_at": datetime.now(timezone.utc),
    })
    response = JSONResponse(content={
        "user": {"user_id": user_id, "full_name": data.full_name, "email": data.email, "role": "user"},
        "session_token": session_token,
    })
    response.set_cookie("session_token", session_token, path="/", httponly=True, secure=True, samesite="none", max_age=7*24*3600)
    return response

@api_router.post("/auth/user-login")
async def user_login(credentials: UserLogin):
    # Check app_users first
    user = await db.app_users.find_one({"email": credentials.email}, {"_id": 0})
    is_admin = False
    if not user:
        # Also check admin users collection
        admin = await db.users.find_one({"email": credentials.email}, {"_id": 0})
        if admin:
            user = admin
            is_admin = True
    if not user:
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    if not user.get("password"):
        raise HTTPException(status_code=401, detail="Bu hesap Google ile oluşturulmuş. Google ile giriş yapın.")
    if not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    # Determine user_id
    user_id = user.get("user_id") or user.get("id") or f"user_{uuid.uuid4().hex[:12]}"
    session_token = f"sess_{uuid.uuid4().hex}"
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
        "created_at": datetime.now(timezone.utc),
    })
    # Update last login & log activity
    await db.app_users.update_one({"user_id": user_id}, {"$set": {"last_login_at": datetime.now(timezone.utc).isoformat()}})
    await db.activity_logs.insert_one({
        "log_id": str(uuid.uuid4()), "user_id": user_id, "action_type": "login",
        "metadata": {"auth_provider": "email", "email": credentials.email},
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    role = user.get("role", "admin" if is_admin else "user")
    response_data = {
        "user": {"user_id": user_id, "full_name": user.get("full_name", ""), "email": user["email"], "role": role},
        "session_token": session_token,
    }
    # For admin users also return JWT access_token
    if is_admin:
        response_data["access_token"] = create_access_token({"sub": credentials.email})
    response = JSONResponse(content=response_data)
    response.set_cookie("session_token", session_token, path="/", httponly=True, secure=True, samesite="none", max_age=7*24*3600)
    return response

# ============= GOOGLE OAUTH =============

EMERGENT_AUTH_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"

@api_router.post("/auth/google-session")
async def google_session(request: Request):
    body = await request.json()
    session_id = body.get("session_id")
    if not session_id:
        raise HTTPException(status_code=400, detail="session_id required")
    # Exchange session_id with Emergent Auth
    logger.info(f"Google session exchange attempt, session_id length: {len(session_id)}")
    try:
        resp = requests.get(EMERGENT_AUTH_URL, headers={"X-Session-ID": session_id}, timeout=10)
        logger.info(f"Emergent Auth response status: {resp.status_code}")
        if resp.status_code != 200:
            logger.error(f"Emergent Auth failed: {resp.status_code} - {resp.text}")
            raise HTTPException(status_code=401, detail=f"Google auth failed: {resp.json().get('detail', {}).get('error_description', 'Unknown error')}")
    except requests.exceptions.RequestException as e:
        logger.error(f"Emergent Auth request error: {e}")
        raise HTTPException(status_code=503, detail="Auth service unavailable")
    google_data = resp.json()
    email = google_data.get("email")
    name = google_data.get("name", "")
    picture = google_data.get("picture", "")
    # Find or create user
    existing = await db.app_users.find_one({"email": email}, {"_id": 0})
    if existing:
        user_id = existing["user_id"]
        await db.app_users.update_one({"user_id": user_id}, {"$set": {"full_name": name, "picture": picture}})
    else:
        user_id = f"user_{uuid.uuid4().hex[:12]}"
        await db.app_users.insert_one({
            "user_id": user_id,
            "full_name": name,
            "email": email,
            "phone": "",
            "password": None,
            "picture": picture,
            "role": "user",
            "auth_provider": "google",
            "created_at": datetime.now(timezone.utc).isoformat(),
        })
    # Create session
    session_token = f"sess_{uuid.uuid4().hex}"
    await db.user_sessions.insert_one({
        "user_id": user_id,
        "session_token": session_token,
        "expires_at": datetime.now(timezone.utc) + timedelta(days=7),
        "created_at": datetime.now(timezone.utc),
    })
    response = JSONResponse(content={
        "user": {"user_id": user_id, "full_name": name, "email": email, "role": "user", "picture": picture},
        "session_token": session_token,
    })
    response.set_cookie("session_token", session_token, path="/", httponly=True, secure=True, samesite="none", max_age=7*24*3600)
    return response

@api_router.post("/auth/logout")
async def logout_user(request: Request):
    token = request.cookies.get("session_token")
    if token:
        await db.user_sessions.delete_one({"session_token": token})
    response = JSONResponse(content={"message": "Logged out"})
    response.delete_cookie("session_token", path="/")
    return response

# ============= CROSS-SITE AUTH (e-ipat.com entegrasyonu) =============

@api_router.post("/auth/cross-site-token")
async def create_cross_site_token(request: Request):
    """Giriş yapmış kullanıcı için 5 dakikalık tek kullanımlık çapraz site token'ı oluşturur."""
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    token = str(uuid.uuid4())
    await db.cross_site_tokens.insert_one({
        "token": token,
        "user_id": user["user_id"],
        "email": user.get("email", ""),
        "full_name": user.get("full_name", ""),
        "phone": user.get("phone", ""),
        "plan": user.get("plan", "free"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(minutes=5)).isoformat(),
        "used": False,
    })
    return {"token": token}

@api_router.get("/auth/verify-cross-site-token")
async def verify_cross_site_token(token: str = Query(...)):
    """e-ipat.com bu endpoint'i çağırarak token'ı doğrular ve kullanıcı bilgilerini alır."""
    doc = await db.cross_site_tokens.find_one({"token": token}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=401, detail="Geçersiz token")
    if doc.get("used"):
        raise HTTPException(status_code=401, detail="Token zaten kullanıldı")
    expires_at = datetime.fromisoformat(doc["expires_at"])
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Token süresi dolmuş")
    # Token'ı kullanıldı olarak işaretle (tek kullanımlık)
    await db.cross_site_tokens.update_one({"token": token}, {"$set": {"used": True}})
    return {
        "valid": True,
        "user_id": doc["user_id"],
        "email": doc["email"],
        "full_name": doc["full_name"],
        "phone": doc.get("phone", ""),
        "plan": doc.get("plan", "free"),
    }

# --- Remote Auth: e-ipat.com bu endpoint'leri kullanarak mrxakademi üyelik veritabanını paylaşır ---

@api_router.post("/auth/remote/login")
async def remote_login(credentials: UserLogin):
    """e-ipat.com login formu bu endpoint'i çağırır. mrxakademi DB'sinde doğrular."""
    user = await db.app_users.find_one({"email": credentials.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    if user.get("auth_provider") == "google":
        raise HTTPException(status_code=401, detail="Bu hesap Google ile oluşturulmuş. Google ile giriş yapın.")
    if not verify_password(credentials.password, user.get("password", "")):
        raise HTTPException(status_code=401, detail="E-posta veya şifre hatalı")
    if user.get("status") == "banned":
        raise HTTPException(status_code=403, detail="Hesabınız askıya alınmıştır")
    session_token = f"sess_{uuid.uuid4().hex}"
    await db.user_sessions.insert_one({
        "session_token": session_token,
        "user_id": user["user_id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "source": "e-ipat.com",
    })
    return {
        "user": {"user_id": user["user_id"], "full_name": user.get("full_name", ""),
                 "email": user["email"], "phone": user.get("phone", ""),
                 "plan": user.get("plan", "free"), "role": user.get("role", "user")},
        "session_token": session_token,
    }

@api_router.post("/auth/remote/register")
async def remote_register(data: UserRegister):
    """e-ipat.com kayıt formu bu endpoint'i çağırır. mrxakademi DB'sine kaydeder."""
    existing = await db.app_users.find_one({"email": data.email})
    if existing:
        raise HTTPException(status_code=409, detail="Bu e-posta zaten kayıtlı")
    user_id = f"user_{uuid.uuid4().hex[:12]}"
    new_user = {
        "user_id": user_id, "full_name": data.full_name, "phone": data.phone,
        "email": data.email, "password": hash_password(data.password),
        "role": "user", "plan": "free", "status": "active",
        "auth_provider": "email", "registered_from": "e-ipat.com",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.app_users.insert_one(new_user)
    new_user.pop("_id", None)
    new_user.pop("password", None)
    session_token = f"sess_{uuid.uuid4().hex}"
    await db.user_sessions.insert_one({
        "session_token": session_token, "user_id": user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": (datetime.now(timezone.utc) + timedelta(days=7)).isoformat(),
        "source": "e-ipat.com",
    })
    return {"user": new_user, "session_token": session_token}

@api_router.get("/auth/remote/verify")
async def remote_verify_session(token: str = Query(...)):
    """e-ipat.com session token'ı doğrulamak için bu endpoint'i çağırır."""
    session = await db.user_sessions.find_one({"session_token": token}, {"_id": 0})
    if not session:
        raise HTTPException(status_code=401, detail="Geçersiz oturum")
    expires_at = session.get("expires_at")
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at and expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if expires_at and expires_at < datetime.now(timezone.utc):
        raise HTTPException(status_code=401, detail="Oturum süresi dolmuş")
    user = await db.app_users.find_one({"user_id": session["user_id"]}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=401, detail="Kullanıcı bulunamadı")
    return {
        "valid": True,
        "user": {"user_id": user["user_id"], "full_name": user.get("full_name", ""),
                 "email": user.get("email", ""), "phone": user.get("phone", ""),
                 "plan": user.get("plan", "free"), "role": user.get("role", "user")}
    }

@api_router.post("/auth/remote/logout")
async def remote_logout(token: str = Query(...)):
    """e-ipat.com çıkış yaparken session'ı sonlandırır."""
    await db.user_sessions.delete_one({"session_token": token})
    return {"message": "Oturum sonlandırıldı"}

# ============= USER MANAGEMENT (ADMIN) =============

@api_router.get("/admin/app-users")
async def list_app_users(
    admin: dict = Depends(require_admin),
    search: str = Query(""), role: str = Query(""),
    plan: str = Query(""), status: str = Query(""),
    page: int = Query(1), limit: int = Query(20)
):
    query = {}
    if search:
        query["$or"] = [
            {"full_name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"user_id": {"$regex": search, "$options": "i"}},
        ]
    if role: query["role"] = role
    if plan: query["membership_plan"] = plan
    if status: query["status"] = status
    total = await db.app_users.count_documents(query)
    skip = (page - 1) * limit
    cursor = db.app_users.find(query, {"_id": 0, "password": 0}).skip(skip).limit(limit).sort("created_at", -1)
    users = await cursor.to_list(length=limit)
    return {"users": users, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@api_router.get("/admin/app-users/{user_id}")
async def get_app_user(user_id: str, admin: dict = Depends(require_admin)):
    user = await db.app_users.find_one({"user_id": user_id}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    return user

@api_router.put("/admin/app-users/{user_id}")
async def update_app_user(user_id: str, data: UserUpdateAdmin, admin: dict = Depends(require_admin)):
    update_data = {k: v for k, v in data.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.app_users.update_one({"user_id": user_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    user = await db.app_users.find_one({"user_id": user_id}, {"_id": 0, "password": 0})
    return user

@api_router.put("/admin/app-users/{user_id}/status")
async def update_user_status(user_id: str, status: str = Query(...), admin: dict = Depends(require_admin)):
    if status not in ["active", "passive", "banned"]:
        raise HTTPException(status_code=400, detail="Geçersiz durum")
    await db.app_users.update_one(
        {"user_id": user_id},
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    await db.activity_logs.insert_one({
        "log_id": str(uuid.uuid4()), "user_id": user_id, "action_type": f"status_changed_{status}",
        "metadata": {"by_admin": admin.get("email"), "new_status": status},
        "created_at": datetime.now(timezone.utc).isoformat(),
    })
    return {"status": status}

@api_router.put("/admin/app-users/{user_id}/membership")
async def update_user_membership(user_id: str, data: MembershipUpdate, admin: dict = Depends(require_admin)):
    user = await db.app_users.find_one({"user_id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if data.membership_plan is not None: update_data["membership_plan"] = data.membership_plan
    if data.membership_start_at is not None: update_data["membership_start_at"] = data.membership_start_at
    if data.membership_end_at is not None: update_data["membership_end_at"] = data.membership_end_at
    if data.membership_active is not None: update_data["membership_active"] = data.membership_active
    if data.extend_days:
        current_end = user.get("membership_end_at")
        try:
            end_dt = datetime.fromisoformat(current_end.replace("Z", "+00:00")) if current_end else datetime.now(timezone.utc)
        except Exception:
            end_dt = datetime.now(timezone.utc)
        if end_dt.tzinfo is None: end_dt = end_dt.replace(tzinfo=timezone.utc)
        update_data["membership_end_at"] = (end_dt + timedelta(days=data.extend_days)).isoformat()
    await db.app_users.update_one({"user_id": user_id}, {"$set": update_data})
    updated = await db.app_users.find_one({"user_id": user_id}, {"_id": 0, "password": 0})
    return updated

@api_router.put("/admin/app-users/{user_id}/permissions")
async def update_user_permissions(user_id: str, data: PermissionsUpdate, admin: dict = Depends(require_admin)):
    await db.app_users.update_one(
        {"user_id": user_id},
        {"$set": {"permissions": data.permissions, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"permissions": data.permissions}

@api_router.put("/admin/app-users/{user_id}/note")
async def update_user_note(user_id: str, data: NoteUpdate, admin: dict = Depends(require_admin)):
    await db.app_users.update_one(
        {"user_id": user_id},
        {"$set": {"admin_note": data.admin_note, "tags": data.tags, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"admin_note": data.admin_note, "tags": data.tags}

@api_router.get("/admin/app-users/{user_id}/activity")
async def get_user_activity(
    user_id: str, admin: dict = Depends(require_admin),
    page: int = Query(1), limit: int = Query(20)
):
    query = {"user_id": user_id}
    total = await db.activity_logs.count_documents(query)
    skip = (page - 1) * limit
    cursor = db.activity_logs.find(query, {"_id": 0}).skip(skip).limit(limit).sort("created_at", -1)
    logs = await cursor.to_list(length=limit)
    return {"logs": logs, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@api_router.post("/admin/app-users/{user_id}/end-sessions")
async def end_user_sessions(user_id: str, admin: dict = Depends(require_admin)):
    result = await db.user_sessions.delete_many({"user_id": user_id})
    return {"deleted_count": result.deleted_count}

PROJECT_TYPES = ["TOKİ", "Emlak Konut", "Özel Proje", "Kamu Projesi"]

@api_router.post("/admin/projects")
async def create_project(
    admin: dict = Depends(require_admin),
    project_name: str = Form(...),
    city: str = Form(...),
    district: str = Form(...),
    neighborhood: str = Form(""),
    description: str = Form(""),
    project_type: str = Form("TOKİ"),
    total_housing: int = Form(0),
    commercial_count: int = Form(0),
    school_count: int = Form(0),
    mosque_count: int = Form(0),
    social_facility_count: int = Form(0),
    project_area_sqm: float = Form(0),
    start_date: str = Form(""),
    planned_end_date: str = Form(""),
    progress_percentage: int = Form(0),
    location_lat: float = Form(41.0082),
    location_lng: float = Form(28.9784),
):
    project = {
        "id": str(uuid.uuid4()),
        "project_name": project_name,
        "city": city,
        "district": district,
        "neighborhood": neighborhood,
        "description": description,
        "project_type": project_type,
        "total_housing": total_housing,
        "commercial_count": commercial_count,
        "school_count": school_count,
        "mosque_count": mosque_count,
        "social_facility_count": social_facility_count,
        "project_area_sqm": project_area_sqm,
        "start_date": start_date,
        "planned_end_date": planned_end_date,
        "progress_percentage": progress_percentage,
        "location": {"lat": location_lat, "lng": location_lng},
        "youtube_videos": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.projects.insert_one(project)
    project.pop("_id", None)
    return project

@api_router.get("/projects")
async def get_projects(city: Optional[str] = None, district: Optional[str] = None, project_name: Optional[str] = None, project_type: Optional[str] = None):
    query = {}
    if city:
        query["city"] = {"$regex": city, "$options": "i"}
    if district:
        query["district"] = {"$regex": district, "$options": "i"}
    if project_name:
        query["project_name"] = {"$regex": project_name, "$options": "i"}
    if project_type:
        query["project_type"] = project_type
    projects = await db.projects.find(query, {"_id": 0}).limit(500).to_list(500)
    return projects

@api_router.get("/projects/{project_id}")
async def get_project(project_id: str):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project

@api_router.put("/admin/projects/{project_id}")
async def update_project(
    project_id: str,
    admin: dict = Depends(require_admin),
    project_name: str = Form(...),
    city: str = Form(...),
    district: str = Form(...),
    neighborhood: str = Form(""),
    description: str = Form(""),
    project_type: str = Form("TOKİ"),
    total_housing: int = Form(0),
    commercial_count: int = Form(0),
    school_count: int = Form(0),
    mosque_count: int = Form(0),
    social_facility_count: int = Form(0),
    project_area_sqm: float = Form(0),
    start_date: str = Form(""),
    planned_end_date: str = Form(""),
    progress_percentage: int = Form(0),
    location_lat: float = Form(41.0082),
    location_lng: float = Form(28.9784),
):
    update_data = {
        "project_name": project_name, "city": city, "district": district,
        "neighborhood": neighborhood, "description": description, "project_type": project_type,
        "total_housing": total_housing, "commercial_count": commercial_count,
        "school_count": school_count, "mosque_count": mosque_count,
        "social_facility_count": social_facility_count, "project_area_sqm": project_area_sqm,
        "start_date": start_date, "planned_end_date": planned_end_date,
        "progress_percentage": progress_percentage,
        "location": {"lat": location_lat, "lng": location_lng},
    }
    result = await db.projects.update_one({"id": project_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    return project

@api_router.delete("/admin/projects/{project_id}")
async def delete_project(project_id: str, admin: dict = Depends(require_admin)):
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    # Clean up related data
    await db.project_adas.delete_many({"project_id": project_id})
    await db.project_parsels.delete_many({"project_id": project_id})
    await db.project_media.delete_many({"project_id": project_id})
    await db.project_documents.delete_many({"project_id": project_id})
    await db.project_map_layers.delete_many({"project_id": project_id})
    return {"message": "Project deleted"}

# ============= ADA (BLOCK) CRUD =============

@api_router.post("/admin/projects/{project_id}/adas")
async def create_ada(project_id: str, admin: dict = Depends(require_admin), ada_no: str = Form(...), description: str = Form("")):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    existing = await db.project_adas.find_one({"project_id": project_id, "ada_no": ada_no})
    if existing:
        raise HTTPException(status_code=400, detail=f"Ada {ada_no} already exists in this project")
    ada = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "ada_no": ada_no,
        "description": description,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_adas.insert_one(ada)
    ada.pop("_id", None)
    return ada

@api_router.get("/projects/{project_id}/adas")
async def get_adas(project_id: str):
    adas = await db.project_adas.find({"project_id": project_id}, {"_id": 0}).sort("ada_no", 1).limit(500).to_list(500)
    return adas

@api_router.put("/admin/adas/{ada_id}")
async def update_ada(ada_id: str, admin: dict = Depends(require_admin), ada_no: str = Form(...), description: str = Form("")):
    result = await db.project_adas.update_one({"id": ada_id}, {"$set": {"ada_no": ada_no, "description": description}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ada not found")
    ada = await db.project_adas.find_one({"id": ada_id}, {"_id": 0})
    return ada

@api_router.delete("/admin/adas/{ada_id}")
async def delete_ada(ada_id: str, admin: dict = Depends(require_admin)):
    ada = await db.project_adas.find_one({"id": ada_id})
    if not ada:
        raise HTTPException(status_code=404, detail="Ada not found")
    await db.project_adas.delete_one({"id": ada_id})
    await db.project_parsels.delete_many({"ada_id": ada_id})
    return {"message": "Ada and its parcels deleted"}

# ============= PARSEL (PARCEL) CRUD =============

@api_router.post("/admin/adas/{ada_id}/parsels")
async def create_parsel(ada_id: str, admin: dict = Depends(require_admin), parsel_no: str = Form(...), area_sqm: float = Form(0), note: str = Form("")):
    ada = await db.project_adas.find_one({"id": ada_id})
    if not ada:
        raise HTTPException(status_code=404, detail="Ada not found")
    parsel = {
        "id": str(uuid.uuid4()),
        "project_id": ada["project_id"],
        "ada_id": ada_id,
        "parsel_no": parsel_no,
        "area_sqm": area_sqm if area_sqm > 0 else None,
        "note": note,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_parsels.insert_one(parsel)
    parsel.pop("_id", None)
    return parsel

@api_router.get("/projects/{project_id}/parsels")
async def get_parsels(project_id: str):
    parsels = await db.project_parsels.find({"project_id": project_id}, {"_id": 0}).limit(5000).to_list(5000)
    return parsels

@api_router.put("/admin/parsels/{parsel_id}")
async def update_parsel(parsel_id: str, admin: dict = Depends(require_admin), parsel_no: str = Form(...), area_sqm: float = Form(0), note: str = Form("")):
    result = await db.project_parsels.update_one({"id": parsel_id}, {"$set": {"parsel_no": parsel_no, "area_sqm": area_sqm if area_sqm > 0 else None, "note": note}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Parsel not found")
    parsel = await db.project_parsels.find_one({"id": parsel_id}, {"_id": 0})
    return parsel

@api_router.delete("/admin/parsels/{parsel_id}")
async def delete_parsel(parsel_id: str, admin: dict = Depends(require_admin)):
    result = await db.project_parsels.delete_one({"id": parsel_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Parsel not found")
    return {"message": "Parsel deleted"}

# ============= EXCEL IMPORT =============

@api_router.post("/admin/projects/{project_id}/import-excel")
async def import_excel(project_id: str, file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Only xlsx, xls, csv files are supported")

    content = await file.read()
    errors = []
    success_count = 0

    try:
        if ext == "csv":
            import csv
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                rows.append(row_dict)

        for idx, row in enumerate(rows, start=2):
            ada_no = str(row.get("Ada", "") or "").strip()
            parsel_no = str(row.get("Parsel", "") or "").strip()
            area_str = row.get("Alan_m2", "")
            note = str(row.get("Not", "") or "").strip()

            if not ada_no or not parsel_no:
                errors.append({"row": idx, "error": "Ada veya Parsel boş"})
                continue

            area_sqm = None
            if area_str:
                try:
                    area_sqm = float(area_str)
                except (ValueError, TypeError):
                    errors.append({"row": idx, "error": f"Geçersiz alan değeri: {area_str}"})
                    continue

            # Find or create ada
            ada = await db.project_adas.find_one({"project_id": project_id, "ada_no": ada_no})
            if not ada:
                ada = {
                    "id": str(uuid.uuid4()),
                    "project_id": project_id,
                    "ada_no": ada_no,
                    "description": "",
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                await db.project_adas.insert_one(ada)

            ada_id = ada["id"]
            parsel = {
                "id": str(uuid.uuid4()),
                "project_id": project_id,
                "ada_id": ada_id,
                "parsel_no": parsel_no,
                "area_sqm": area_sqm,
                "note": note,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.project_parsels.insert_one(parsel)
            success_count += 1

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File processing error: {str(e)}")

    return {"success_count": success_count, "error_count": len(errors), "errors": errors}

@api_router.get("/admin/excel-template")
async def download_excel_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ada-Parsel"
    ws.append(["Ada", "Parsel", "Alan_m2", "Not"])
    ws.append(["33", "1", "500", "Örnek not"])
    ws.append(["33", "2", "750", ""])
    ws.append(["34", "7", "1200", "Köşe parsel"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ada_parsel_sablonu.xlsx"}
    )

# ============= PROJECT EXCEL IMPORT =============

@api_router.get("/admin/project-excel-template")
async def download_project_excel_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projeler"
    headers = ["Proje_Adi", "Il", "Ilce", "Mahalle", "Proje_Tipi", "Aciklama",
               "Konut_Sayisi", "Ticari_Alan", "Okul", "Cami", "Sosyal_Tesis",
               "Proje_Alani_m2", "Baslangic_Tarihi", "Bitis_Tarihi", "Ilerleme_Yuzde",
               "Enlem", "Boylam"]
    ws.append(headers)
    ws.append(["Örnek Proje", "İstanbul", "Arnavutköy", "Tayakadın", "TOKİ", "Açıklama metni",
               "761", "197", "1", "1", "2", "50000", "2024-01-15", "2026-06-30", "45",
               "41.0082", "28.9784"])
    # Adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=proje_sablonu.xlsx"}
    )

@api_router.post("/admin/projects/import-excel")
async def import_projects_excel(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Sadece xlsx, xls, csv desteklenir")

    content = await file.read()
    errors = []
    success_count = 0
    created_ids = []

    try:
        if ext == "csv":
            import csv
            reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            headers = [str(cell.value or "").strip() for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = val
                rows.append(row_dict)

        for idx, row in enumerate(rows, start=2):
            name = str(row.get("Proje_Adi", "") or "").strip()
            city = str(row.get("Il", "") or "").strip()
            district = str(row.get("Ilce", "") or "").strip()

            if not name or not city:
                errors.append({"row": idx, "error": "Proje_Adi veya Il boş"})
                continue

            def safe_int(val, default=0):
                try:
                    return int(float(val)) if val else default
                except (ValueError, TypeError):
                    return default

            def safe_float(val, default=0.0):
                try:
                    return float(val) if val else default
                except (ValueError, TypeError):
                    return default

            project = {
                "id": str(uuid.uuid4()),
                "project_name": name,
                "city": city,
                "district": district,
                "neighborhood": str(row.get("Mahalle", "") or "").strip(),
                "description": str(row.get("Aciklama", "") or "").strip(),
                "project_type": str(row.get("Proje_Tipi", "TOKİ") or "TOKİ").strip(),
                "total_housing": safe_int(row.get("Konut_Sayisi")),
                "commercial_count": safe_int(row.get("Ticari_Alan")),
                "school_count": safe_int(row.get("Okul")),
                "mosque_count": safe_int(row.get("Cami")),
                "social_facility_count": safe_int(row.get("Sosyal_Tesis")),
                "project_area_sqm": safe_float(row.get("Proje_Alani_m2")),
                "start_date": str(row.get("Baslangic_Tarihi", "") or "").strip(),
                "planned_end_date": str(row.get("Bitis_Tarihi", "") or "").strip(),
                "progress_percentage": safe_int(row.get("Ilerleme_Yuzde")),
                "location": {
                    "lat": safe_float(row.get("Enlem"), 41.0082),
                    "lng": safe_float(row.get("Boylam"), 28.9784),
                },
                "youtube_videos": [],
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.projects.insert_one(project)
            project.pop("_id", None)
            created_ids.append(project["id"])
            success_count += 1

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Dosya işleme hatası: {str(e)}")

    return {"success_count": success_count, "error_count": len(errors), "errors": errors, "created_ids": created_ids}

# ============= MEDIA MANAGEMENT =============

MEDIA_CATEGORIES = ["Altyapı", "Blok Resimleri", "Peyzaj", "Zemin", "Drone", "Master Plan"]

@api_router.post("/admin/projects/{project_id}/media")
async def upload_media(project_id: str, file: UploadFile = File(...), category: str = Form(...), admin: dict = Depends(require_admin)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "bin"
    storage_path = f"{APP_NAME}/media/{project_id}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    content_type = file.content_type or MIME_TYPES.get(ext, "application/octet-stream")
    put_object(storage_path, data, content_type)

    media = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "category": category,
        "storage_path": storage_path,
        "original_filename": file.filename,
        "content_type": content_type,
        "size": len(data),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_media.insert_one(media)
    media.pop("_id", None)
    return media

@api_router.get("/projects/{project_id}/media")
async def get_project_media(project_id: str, category: Optional[str] = None):
    query = {"project_id": project_id}
    if category:
        query["category"] = category
    media = await db.project_media.find(query, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)
    return media

@api_router.delete("/admin/media/{media_id}")
async def delete_media(media_id: str, admin: dict = Depends(require_admin)):
    result = await db.project_media.delete_one({"id": media_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Media not found")
    return {"message": "Media deleted"}

# ============= DOCUMENT MANAGEMENT =============

DOC_TYPES = ["Zemin Etüt", "Jeoloji / Jeoteknik", "ÇED", "İhale Belgeleri", "Plan Notları", "Vaziyet Planı", "Diğer"]

@api_router.post("/admin/projects/{project_id}/documents")
async def upload_document(project_id: str, file: UploadFile = File(...), title: str = Form(...), doc_type: str = Form(...), admin: dict = Depends(require_admin)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "bin"
    storage_path = f"{APP_NAME}/documents/{project_id}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    content_type = file.content_type or MIME_TYPES.get(ext, "application/octet-stream")
    put_object(storage_path, data, content_type)

    doc = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "title": title,
        "doc_type": doc_type,
        "storage_path": storage_path,
        "original_filename": file.filename,
        "content_type": content_type,
        "size": len(data),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_documents.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/projects/{project_id}/documents")
async def get_project_documents(project_id: str):
    docs = await db.project_documents.find({"project_id": project_id}, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)
    return docs

@api_router.delete("/admin/documents/{doc_id}")
async def delete_document(doc_id: str, admin: dict = Depends(require_admin)):
    result = await db.project_documents.delete_one({"id": doc_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": "Document deleted"}

# ============= VIDEO MANAGEMENT =============

@api_router.post("/admin/projects/{project_id}/videos")
async def add_video(project_id: str, admin: dict = Depends(require_admin), youtube_url: str = Form(...)):
    result = await db.projects.update_one({"id": project_id}, {"$push": {"youtube_videos": youtube_url}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"message": "Video added"}

@api_router.delete("/admin/projects/{project_id}/videos")
async def remove_video(project_id: str, admin: dict = Depends(require_admin), youtube_url: str = Form(...)):
    result = await db.projects.update_one({"id": project_id}, {"$pull": {"youtube_videos": youtube_url}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"message": "Video removed"}

# ============= MAP LAYERS =============

@api_router.post("/admin/projects/{project_id}/map-layers")
async def upload_map_layer(project_id: str, file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    if ext not in ("kml", "kmz", "geojson", "json"):
        raise HTTPException(status_code=400, detail="Only KML, KMZ, GeoJSON files are supported")

    storage_path = f"{APP_NAME}/map-layers/{project_id}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    content_type = file.content_type or MIME_TYPES.get(ext, "application/octet-stream")
    put_object(storage_path, data, content_type)

    layer = {
        "id": str(uuid.uuid4()),
        "project_id": project_id,
        "storage_path": storage_path,
        "original_filename": file.filename,
        "content_type": content_type,
        "file_type": ext.upper(),
        "size": len(data),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_map_layers.insert_one(layer)
    layer.pop("_id", None)

    # Auto-extract centroid and update project location
    centroid = None
    try:
        if ext in ("kml", "kmz"):
            centroid = extract_centroid_from_kml(ext, data)
        elif ext in ("geojson", "json"):
            centroid = extract_centroid_from_geojson(data)
    except Exception:
        pass
    if centroid:
        await db.projects.update_one({"id": project_id}, {"$set": {"location": centroid}})
        layer["location_updated"] = centroid

    return layer

@api_router.get("/projects/{project_id}/map-layers")
async def get_map_layers(project_id: str):
    layers = await db.project_map_layers.find({"project_id": project_id}, {"_id": 0}).to_list(100)
    return layers

@api_router.delete("/admin/map-layers/{layer_id}")
async def delete_map_layer(layer_id: str, admin: dict = Depends(require_admin)):
    result = await db.project_map_layers.delete_one({"id": layer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Map layer not found")
    return {"message": "Map layer deleted"}

# ============= FILE SERVING =============

@api_router.get("/files/{path:path}")
async def serve_file(path: str):
    try:
        data, content_type = get_object(path)
        return Response(content=data, media_type=content_type)
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")

# ============= MAP LAYER DATA (GeoJSON content) =============

@api_router.get("/projects/{project_id}/map-layers/{layer_id}/data")
async def get_map_layer_data(project_id: str, layer_id: str):
    layer = await db.project_map_layers.find_one({"id": layer_id, "project_id": project_id}, {"_id": 0})
    if not layer:
        raise HTTPException(status_code=404, detail="Map layer not found")
    try:
        data, _ = get_object(layer["storage_path"])
        if layer["file_type"] in ("GEOJSON", "JSON"):
            return json.loads(data)
        # KMZ: unzip and return inner KML
        if layer["file_type"] == "KMZ":
            import zipfile
            import io
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                kml_files = [n for n in z.namelist() if n.endswith('.kml')]
                if not kml_files:
                    raise HTTPException(status_code=400, detail="No KML found inside KMZ")
                kml_content = z.read(kml_files[0])
            return Response(content=kml_content, media_type="application/vnd.google-earth.kml+xml")
        return Response(content=data, media_type=layer["content_type"])
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot read layer: {str(e)}")

# ============= BACKWARD COMPAT: TOKI ENDPOINTS =============

@api_router.get("/toki/projects")
async def get_toki_projects(city: Optional[str] = None, district: Optional[str] = None, project_name: Optional[str] = None):
    query = {}
    if city:
        query["city"] = {"$regex": city, "$options": "i"}
    if district:
        query["district"] = {"$regex": district, "$options": "i"}
    if project_name:
        query["project_name"] = {"$regex": project_name, "$options": "i"}
    return await db.projects.find(query, {"_id": 0}).limit(500).to_list(500)

@api_router.get("/toki/projects/{project_id}")
async def get_toki_project(project_id: str):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project

# ============= INVESTMENT CALCULATOR =============

@api_router.post("/investment/calculate", response_model=InvestmentResult)
async def calculate_investment(data: InvestmentCalculation):
    total_construction_area = data.land_size_sqm * data.emsal
    estimated_apartments = int(total_construction_area / 100)
    total_construction_cost = total_construction_area * data.construction_cost_per_sqm
    estimated_project_value = total_construction_cost * 1.5
    potential_profit = estimated_project_value - total_construction_cost
    roi_percentage = (potential_profit / total_construction_cost) * 100 if total_construction_cost > 0 else 0
    return InvestmentResult(
        total_construction_area=total_construction_area,
        estimated_apartments=estimated_apartments,
        total_construction_cost=total_construction_cost,
        estimated_project_value=estimated_project_value,
        potential_profit=potential_profit,
        roi_percentage=round(roi_percentage, 2)
    )

# ============= MEGA PROJECTS ADMIN =============

@api_router.post("/admin/mega-projects")
async def create_mega_project(
    admin: dict = Depends(require_admin),
    name: str = Form(...), category: str = Form(...), description: str = Form(""),
    timeline: str = Form(""), location_lat: float = Form(41.0082), location_lng: float = Form(28.9784),
):
    project = {
        "id": str(uuid.uuid4()), "name": name, "category": category, "description": description,
        "timeline": timeline, "location": {"lat": location_lat, "lng": location_lng},
        "images": [], "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.mega_projects.insert_one(project)
    project.pop("_id", None)
    return project

@api_router.get("/mega-projects")
async def get_mega_projects():
    projects = await db.mega_projects.find({}, {"_id": 0}).limit(500).to_list(500)
    # Also include all projects from projects collection (auto-map)
    all_projects = await db.projects.find({}, {"_id": 0, "id": 1, "project_name": 1, "city": 1, "district": 1, "project_type": 1, "location": 1, "progress_percentage": 1}).limit(500).to_list(500)
    for p in all_projects:
        if p.get("location"):
            projects.append({
                "id": p["id"], "name": p.get("project_name", ""), "category": p.get("project_type", "TOKİ"),
                "description": f"{p.get('city','')} / {p.get('district','')}", "timeline": "",
                "location": p["location"], "images": [], "from_projects": True,
                "progress_percentage": p.get("progress_percentage", 0),
            })
    return projects

@api_router.put("/admin/mega-projects/{project_id}")
async def update_mega_project(
    project_id: str, admin: dict = Depends(require_admin),
    name: str = Form(...), category: str = Form(...), description: str = Form(""),
    timeline: str = Form(""), location_lat: float = Form(41.0082), location_lng: float = Form(28.9784),
):
    result = await db.mega_projects.update_one({"id": project_id}, {"$set": {
        "name": name, "category": category, "description": description,
        "timeline": timeline, "location": {"lat": location_lat, "lng": location_lng},
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    project = await db.mega_projects.find_one({"id": project_id}, {"_id": 0})
    return project

@api_router.delete("/admin/mega-projects/{project_id}")
async def delete_mega_project(project_id: str, admin: dict = Depends(require_admin)):
    result = await db.mega_projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted"}

# ============= e-IPAT (LAND PARCELS) ADMIN =============

@api_router.post("/admin/land-parcels")
async def create_land_parcel(
    admin: dict = Depends(require_admin),
    city: str = Form(...), district: str = Form(...), neighborhood: str = Form(""),
    ada: str = Form(...), parsel: str = Form(...), size_sqm: float = Form(0),
    zoning_info: str = Form(""), development_potential: str = Form(""),
    location_lat: float = Form(41.0082), location_lng: float = Form(28.9784),
):
    parcel = {
        "id": str(uuid.uuid4()), "city": city, "district": district, "neighborhood": neighborhood,
        "ada": ada, "parsel": parsel, "size_sqm": size_sqm, "zoning_info": zoning_info,
        "development_potential": development_potential, "location": {"lat": location_lat, "lng": location_lng},
        "documents": [], "images": [], "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.land_parcels.insert_one(parcel)
    parcel.pop("_id", None)
    return parcel

@api_router.get("/land-parcels")
async def get_land_parcels(city: Optional[str] = None, district: Optional[str] = None):
    query = {}
    if city:
        query["city"] = {"$regex": city, "$options": "i"}
    if district:
        query["district"] = {"$regex": district, "$options": "i"}
    parcels = await db.land_parcels.find(query, {"_id": 0}).limit(500).to_list(500)
    return parcels

@api_router.put("/admin/land-parcels/{parcel_id}")
async def update_land_parcel(
    parcel_id: str, admin: dict = Depends(require_admin),
    city: str = Form(...), district: str = Form(...), neighborhood: str = Form(""),
    ada: str = Form(...), parsel: str = Form(...), size_sqm: float = Form(0),
    zoning_info: str = Form(""), development_potential: str = Form(""),
    location_lat: float = Form(41.0082), location_lng: float = Form(28.9784),
):
    result = await db.land_parcels.update_one({"id": parcel_id}, {"$set": {
        "city": city, "district": district, "neighborhood": neighborhood,
        "ada": ada, "parsel": parsel, "size_sqm": size_sqm, "zoning_info": zoning_info,
        "development_potential": development_potential, "location": {"lat": location_lat, "lng": location_lng},
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    p = await db.land_parcels.find_one({"id": parcel_id}, {"_id": 0})
    return p

@api_router.delete("/admin/land-parcels/{parcel_id}")
async def delete_land_parcel(parcel_id: str, admin: dict = Depends(require_admin)):
    result = await db.land_parcels.delete_one({"id": parcel_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted"}


# ============= SEO SETTINGS =============

SEO_PAGES = [
    {"id": "home",           "page_name": "Ana Sayfa",          "path": "/"},
    {"id": "e-konut",        "page_name": "e-Konut Projeleri",  "path": "/toki"},
    {"id": "mega-projects",  "page_name": "Mega Projeler",      "path": "/mega-projects"},
    {"id": "ipat",           "page_name": "e-İPAT Arsa Analizi","path": "/ipar"},
    {"id": "egitim",         "page_name": "Eğitim Platformu",   "path": "/egitim"},
    {"id": "topluluk",       "page_name": "Topluluk Forumu",    "path": "/topluluk"},
    {"id": "yatirim-fonu",   "page_name": "Yatırım Fonu",       "path": "/yatirim-fonu"},
]

SEO_PAGE_CONTEXTS = {
    "home":          "mrxakademi - Türkiye'nin lider PropTech platformu. TOKİ konut projeleri analizi, mega altyapı projeleri haritası, arazi parsel analizi (e-İPAT), gayrimenkul eğitimleri ve yatırımcı topluluğu.",
    "e-konut":       "e-Konut / TOKİ projeleri sayfası. İstanbul Arnavutköy, Mamak gibi illerdeki TOKİ konut projelerini harita üzerinde inceleyin. Proje ilerlemeleri, konut istatistikleri, sosyal tesisler.",
    "mega-projects": "Türkiye'nin mega altyapı projeleri interaktif haritası. Kanal İstanbul, 3. Köprü, Marmaray, havalimanları, metro hatları, otoyollar, sanayi bölgeleri.",
    "ipat":          "e-İPAT ada parsel sorgulama ve analiz aracı. Türkiye genelinde koordinat bazlı parsel sorgulama, imar durumu, parsel alanı, mülkiyet analizi, yatırım değerlendirmesi.",
    "egitim":        "Gayrimenkul ve yatırım eğitim platformu. Online kurslar, canlı seminerler, sertifika programları. Arsa yatırımı, TOKİ projeler, portföy yönetimi, uzman eğitmenler.",
    "topluluk":      "mrxakademi gayrimenkul yatırımcıları topluluğu. Proje yorumları, yatırım tavsiyeleri, soru-cevap, deneyim paylaşımı, güncel haberler.",
    "yatirim-fonu":  "Kurumsal ve bireysel gayrimenkul yatırım fonu fırsatları. Güvenli portföy yönetimi, yüksek kira getirisi, TOKİ ve mega proje yatırımları, başvuru formu.",
}

@api_router.get("/seo")
async def get_all_seo_public():
    """Public endpoint — returns all SEO settings as {page_id: settings}."""
    docs = await db.seo_settings.find({}, {"_id": 0}).to_list(50)
    return {d["page_id"]: d for d in docs}

@api_router.get("/admin/seo")
async def get_all_seo_admin(admin: dict = Depends(require_admin)):
    docs = await db.seo_settings.find({}, {"_id": 0}).to_list(50)
    existing = {d["page_id"] for d in docs}
    for page in SEO_PAGES:
        if page["id"] not in existing:
            default = {
                "page_id": page["id"], "page_name": page["page_name"], "path": page["path"],
                "title": f"{page['page_name']} | mrxakademi", "description": "",
                "keywords": "", "og_title": f"{page['page_name']} | mrxakademi",
                "og_description": "", "og_image": "", "robots": "index,follow",
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.seo_settings.insert_one(default)
            default.pop("_id", None)
            docs.append(default)
    for d in docs:
        d.pop("_id", None)
    return docs

@api_router.put("/admin/seo/{page_id}")
async def update_seo(page_id: str, body: dict, admin: dict = Depends(require_admin)):
    body.pop("_id", None)
    body["page_id"]    = page_id
    body["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.seo_settings.update_one({"page_id": page_id}, {"$set": body}, upsert=True)
    body.pop("_id", None)
    return body

@api_router.post("/admin/seo/generate/{page_id}")
async def generate_seo_ai(page_id: str, admin: dict = Depends(require_admin)):
    """Use Claude Sonnet to generate optimised Turkish SEO for one page."""
    llm_key = os.environ.get("EMERGENT_LLM_KEY")
    if not llm_key:
        raise HTTPException(status_code=500, detail="LLM key bulunamadı")
    context = SEO_PAGE_CONTEXTS.get(page_id, f"mrxakademi {page_id} sayfası")
    prompt = f"""Aşağıdaki sayfa için Türkçe, Google'da üst sıralara çıkacak SEO metaları üret.
Sayfa bağlamı: {context}
Site alan adı: mrxozdemir.com / mrxakademi.com

Kurallar:
- title: 55-60 karakter, en güçlü anahtar kelimeler öne, "| mrxakademi" ile bitsin
- description: 150-160 karakter, net fayda + eylem çağrısı içersin, anahtar kelimeler geçsin
- keywords: 10-12 Türkçe anahtar kelime, virgülle ayrılmış
- og_title: 60 karakter max, sosyal medya paylaşım başlığı, çekici
- og_description: 90-110 karakter, sosyal medya için özet

SADECE geçerli JSON döndür (başka hiçbir şey yazma):
{{
  "title": "",
  "description": "",
  "keywords": "",
  "og_title": "",
  "og_description": ""
}}"""
    try:
        from emergentintegrations.llm.chat import LlmChat, UserMessage as LlmUM
        chat = LlmChat(
            api_key=llm_key,
            session_id=f"seo_{page_id}_{uuid.uuid4()}",
            system_message="Sen bir SEO uzmanısın. Sadece geçerli JSON döndür."
        ).with_model("anthropic", "claude-sonnet-4-5-20250929")
        raw = await chat.send_message(LlmUM(text=prompt))
        raw = raw.strip()
        if "```json" in raw:
            raw = raw.split("```json")[1].split("```")[0].strip()
        elif "```" in raw:
            raw = raw.split("```")[1].split("```")[0].strip()
        return json.loads(raw)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI SEO üretim hatası: {str(e)}")


# ============= EDUCATION SYSTEM =============

class SeminarCreate(BaseModel):
    title: str
    description: str = ""
    date: str = ""
    time: str = ""
    duration: str = ""
    speaker: str = ""
    seminar_type: str = "free"
    location: str = ""
    zoom_link: str = ""
    cover_image: str = ""
    status: str = "active"

class SeminarRegistrationCreate(BaseModel):
    name: str
    phone: str
    email: str

class CourseCreate(BaseModel):
    title: str
    short_description: str = ""
    full_description: str = ""
    cover_image: str = ""
    promo_video: str = ""
    price: float = 0
    discount_price: Optional[float] = None
    level: str = "başlangıç"
    tags: List[str] = []
    status: str = "active"
    order: int = 0
    student_count: int = 0
    rating: float = 5.0

class CourseModuleCreate(BaseModel):
    title: str
    order: int = 0

class CourseLessonCreate(BaseModel):
    title: str
    video_url: str = ""
    pdf_files: List[str] = []
    duration: str = ""
    is_preview: bool = False
    order: int = 0

class LiveTrainingUpdate(BaseModel):
    title: str = "Haftalık Canlı Online Eğitim"
    description: str = ""
    day_of_week: str = ""
    time: str = ""
    zoom_link: str = ""
    status: str = "active"

class LiveArchiveCreate(BaseModel):
    title: str
    video_url: str = ""
    date: str = ""
    thumbnail: str = ""

# --- Public Education Endpoints ---

@api_router.get("/education/courses")
async def list_edu_courses():
    return await db.courses.find({"status": "active"}, {"_id": 0}).sort("order", 1).limit(500).to_list(500)

@api_router.get("/courses")  # backward compat
async def get_courses():
    return await db.courses.find({}, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)

@api_router.get("/education/courses/{course_id}")
async def get_edu_course(course_id: str):
    c = await db.courses.find_one({"id": course_id}, {"_id": 0})
    if not c: raise HTTPException(status_code=404, detail="Kurs bulunamadı")
    return c

@api_router.get("/education/seminars")
async def list_edu_seminars():
    return await db.seminars.find({"status": "active"}, {"_id": 0}).sort("date", -1).limit(500).to_list(500)

@api_router.get("/seminars")  # backward compat
async def get_seminars():
    return await db.seminars.find({}, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)

@api_router.post("/education/seminars/{seminar_id}/register")
async def register_for_seminar(seminar_id: str, data: SeminarRegistrationCreate):
    seminar = await db.seminars.find_one({"id": seminar_id})
    if not seminar: raise HTTPException(status_code=404, detail="Seminer bulunamadı")
    existing = await db.seminar_registrations.find_one({"seminar_id": seminar_id, "email": data.email})
    if existing: raise HTTPException(status_code=400, detail="Bu email ile zaten kayıt yapılmış")
    reg = {
        "id": str(uuid.uuid4()), "seminar_id": seminar_id,
        "seminar_title": seminar.get("title", ""),
        "name": data.name, "phone": data.phone, "email": data.email,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.seminar_registrations.insert_one(reg)
    reg.pop("_id", None)
    return {"message": "Kaydınız başarıyla alındı. Seminer bilgileri size SMS ve email ile gönderilecektir.", "registration": reg}

@api_router.get("/education/live")
async def get_live_training():
    live = await db.weekly_live.find_one({}, {"_id": 0})
    return live or {"title": "Haftalık Canlı Online Eğitim", "description": "Her hafta yatırımcılarla birlikte canlı analiz yapılır.", "day_of_week": "Çarşamba", "time": "20:00", "zoom_link": "", "status": "active", "archives": []}

@api_router.get("/education/page-settings")
async def get_edu_page_settings():
    return await db.edu_page_settings.find_one({}, {"_id": 0}) or {}

# User dashboard
@api_router.get("/user/my-seminars")
async def get_my_seminars(request: Request):
    session = await get_session_user(request)
    if not session: raise HTTPException(status_code=401, detail="Giriş gerekli")
    regs = await db.seminar_registrations.find({"email": session["email"]}, {"_id": 0}).to_list(100)
    return regs

# --- Admin Education Endpoints ---

@api_router.get("/admin/education/courses")
async def admin_list_courses(admin: dict = Depends(require_admin)):
    return await db.courses.find({}, {"_id": 0}).sort("order", 1).limit(500).to_list(500)

@api_router.post("/admin/education/courses")
async def admin_create_course(data: CourseCreate, admin: dict = Depends(require_admin)):
    c = {"id": str(uuid.uuid4()), "modules": [], "created_at": datetime.now(timezone.utc).isoformat(), **data.dict()}
    await db.courses.insert_one(c)
    c.pop("_id", None)
    return c

@api_router.put("/admin/education/courses/{course_id}")
async def admin_update_course(course_id: str, data: CourseCreate, admin: dict = Depends(require_admin)):
    upd = {k: v for k, v in data.dict().items()}
    upd["updated_at"] = datetime.now(timezone.utc).isoformat()
    result = await db.courses.update_one({"id": course_id}, {"$set": upd})
    if result.matched_count == 0: raise HTTPException(status_code=404, detail="Kurs bulunamadı")
    return await db.courses.find_one({"id": course_id}, {"_id": 0})

@api_router.delete("/admin/education/courses/{course_id}")
async def admin_delete_course(course_id: str, admin: dict = Depends(require_admin)):
    await db.courses.delete_one({"id": course_id})
    return {"message": "Silindi"}

@api_router.post("/admin/education/courses/{course_id}/modules")
async def admin_add_module(course_id: str, data: CourseModuleCreate, admin: dict = Depends(require_admin)):
    module = {"module_id": str(uuid.uuid4()), "title": data.title, "order": data.order, "lessons": []}
    result = await db.courses.update_one({"id": course_id}, {"$push": {"modules": module}})
    if result.matched_count == 0: raise HTTPException(status_code=404, detail="Kurs bulunamadı")
    return module

@api_router.put("/admin/education/courses/{course_id}/modules/{module_id}")
async def admin_update_module(course_id: str, module_id: str, data: CourseModuleCreate, admin: dict = Depends(require_admin)):
    await db.courses.update_one(
        {"id": course_id, "modules.module_id": module_id},
        {"$set": {"modules.$.title": data.title, "modules.$.order": data.order}}
    )
    return {"message": "Güncellendi"}

@api_router.delete("/admin/education/courses/{course_id}/modules/{module_id}")
async def admin_delete_module(course_id: str, module_id: str, admin: dict = Depends(require_admin)):
    await db.courses.update_one({"id": course_id}, {"$pull": {"modules": {"module_id": module_id}}})
    return {"message": "Silindi"}

@api_router.post("/admin/education/courses/{course_id}/modules/{module_id}/lessons")
async def admin_add_lesson(course_id: str, module_id: str, data: CourseLessonCreate, admin: dict = Depends(require_admin)):
    lesson = {"lesson_id": str(uuid.uuid4()), **data.dict()}
    await db.courses.update_one(
        {"id": course_id, "modules.module_id": module_id},
        {"$push": {"modules.$.lessons": lesson}}
    )
    return lesson

@api_router.delete("/admin/education/courses/{course_id}/modules/{module_id}/lessons/{lesson_id}")
async def admin_delete_lesson(course_id: str, module_id: str, lesson_id: str, admin: dict = Depends(require_admin)):
    await db.courses.update_one(
        {"id": course_id, "modules.module_id": module_id},
        {"$pull": {"modules.$.lessons": {"lesson_id": lesson_id}}}
    )
    return {"message": "Silindi"}

# --- Admin Seminar Endpoints ---

@api_router.get("/admin/education/seminars")
async def admin_list_seminars(admin: dict = Depends(require_admin)):
    pipeline = [
        {"$sort": {"date": -1}},
        {"$limit": 500},
        {"$lookup": {"from": "seminar_registrations", "localField": "id", "foreignField": "seminar_id", "as": "_regs"}},
        {"$addFields": {"registration_count": {"$size": "$_regs"}}},
        {"$project": {"_id": 0, "_regs": 0}},
    ]
    return await db.seminars.aggregate(pipeline).to_list(500)

@api_router.post("/admin/education/seminars")
async def admin_create_seminar(data: SeminarCreate, admin: dict = Depends(require_admin)):
    s = {"id": str(uuid.uuid4()), "created_at": datetime.now(timezone.utc).isoformat(), **data.dict()}
    await db.seminars.insert_one(s)
    s.pop("_id", None)
    return s

@api_router.put("/admin/education/seminars/{seminar_id}")
async def admin_update_seminar(seminar_id: str, data: SeminarCreate, admin: dict = Depends(require_admin)):
    upd = {**data.dict(), "updated_at": datetime.now(timezone.utc).isoformat()}
    result = await db.seminars.update_one({"id": seminar_id}, {"$set": upd})
    if result.matched_count == 0: raise HTTPException(status_code=404, detail="Seminer bulunamadı")
    return await db.seminars.find_one({"id": seminar_id}, {"_id": 0})

@api_router.delete("/admin/education/seminars/{seminar_id}")
async def admin_delete_seminar(seminar_id: str, admin: dict = Depends(require_admin)):
    await db.seminars.delete_one({"id": seminar_id})
    return {"message": "Silindi"}

@api_router.get("/admin/education/seminars/{seminar_id}/registrations")
async def admin_seminar_registrations(seminar_id: str, admin: dict = Depends(require_admin)):
    return await db.seminar_registrations.find({"seminar_id": seminar_id}, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)

# --- Admin Live Training ---

@api_router.get("/admin/education/live")
async def admin_get_live(admin: dict = Depends(require_admin)):
    live = await db.weekly_live.find_one({}, {"_id": 0})
    return live or {}

@api_router.put("/admin/education/live")
async def admin_update_live(data: LiveTrainingUpdate, admin: dict = Depends(require_admin)):
    upd = {**data.dict(), "updated_at": datetime.now(timezone.utc).isoformat()}
    existing = await db.weekly_live.find_one({})
    if existing:
        await db.weekly_live.update_one({}, {"$set": upd})
    else:
        await db.weekly_live.insert_one({"id": str(uuid.uuid4()), "archives": [], **upd})
    return await db.weekly_live.find_one({}, {"_id": 0})

@api_router.post("/admin/education/live/archives")
async def admin_add_archive(data: LiveArchiveCreate, admin: dict = Depends(require_admin)):
    archive = {"archive_id": str(uuid.uuid4()), **data.dict()}
    existing = await db.weekly_live.find_one({})
    if existing:
        await db.weekly_live.update_one({}, {"$push": {"archives": archive}})
    return archive

@api_router.delete("/admin/education/live/archives/{archive_id}")
async def admin_delete_archive(archive_id: str, admin: dict = Depends(require_admin)):
    await db.weekly_live.update_one({}, {"$pull": {"archives": {"archive_id": archive_id}}})
    return {"message": "Silindi"}

# --- Admin Media Library ---

@api_router.post("/admin/education/media/upload")
async def upload_edu_media(file: UploadFile = File(...), folder: str = Form("genel"), admin: dict = Depends(require_admin)):
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "bin"
    storage_path = f"{APP_NAME}/education/media/{folder}/{uuid.uuid4()}.{ext}"
    data_bytes = await file.read()
    content_type = file.content_type or MIME_TYPES.get(ext, "application/octet-stream")
    put_object(storage_path, data_bytes, content_type)
    media_type = "image" if ext in ["jpg", "jpeg", "png", "gif", "webp"] else "video" if ext in ["mp4", "mov", "avi"] else "pdf" if ext == "pdf" else "other"
    record = {
        "id": str(uuid.uuid4()), "name": file.filename, "original_name": file.filename,
        "type": media_type, "ext": ext, "storage_path": storage_path,
        "size": len(data_bytes), "folder": folder,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.education_media.insert_one(record)
    record.pop("_id", None)
    return record

@api_router.get("/admin/education/media")
async def list_edu_media(folder: str = Query(""), search: str = Query(""), admin: dict = Depends(require_admin)):
    query = {}
    if folder: query["folder"] = folder
    if search: query["name"] = {"$regex": search, "$options": "i"}
    return await db.education_media.find(query, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)

@api_router.delete("/admin/education/media/{media_id}")
async def delete_edu_media(media_id: str, admin: dict = Depends(require_admin)):
    result = await db.education_media.delete_one({"id": media_id})
    if result.deleted_count == 0: raise HTTPException(status_code=404, detail="Medya bulunamadı")
    return {"message": "Silindi"}

# --- Admin Page Settings ---

@api_router.put("/admin/education/page-settings")
async def update_edu_page_settings(sections: List[Dict[str, Any]] = Body(...), admin: dict = Depends(require_admin)):
    await db.edu_page_settings.delete_many({})
    await db.edu_page_settings.insert_one({"sections": sections})
    return {"sections": sections}

# Backward compat - old admin endpoints
@api_router.post("/admin/courses")
async def create_course_compat(admin: dict = Depends(require_admin), title: str = Form(...), description: str = Form(""), video_url: str = Form(""), duration_minutes: int = Form(0), thumbnail: str = Form("")):
    c = {"id": str(uuid.uuid4()), "title": title, "short_description": description, "full_description": "", "cover_image": thumbnail, "promo_video": video_url, "price": 0, "level": "başlangıç", "tags": [], "status": "active", "order": 0, "modules": [], "student_count": duration_minutes, "rating": 5.0, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.courses.insert_one(c)
    c.pop("_id", None)
    return c

@api_router.delete("/admin/courses/{course_id}")
async def delete_course_compat(course_id: str, admin: dict = Depends(require_admin)):
    await db.courses.delete_one({"id": course_id})
    return {"message": "Deleted"}

@api_router.post("/admin/seminars")
async def create_seminar_compat(admin: dict = Depends(require_admin), title: str = Form(...), description: str = Form(""), speaker: str = Form(""), date: str = Form(""), registration_link: str = Form(""), thumbnail: str = Form("")):
    s = {"id": str(uuid.uuid4()), "title": title, "description": description, "speaker": speaker, "date": date, "zoom_link": registration_link, "cover_image": thumbnail, "seminar_type": "free", "time": "", "duration": "", "location": "", "status": "active", "created_at": datetime.now(timezone.utc).isoformat()}
    await db.seminars.insert_one(s)
    s.pop("_id", None)
    return s

@api_router.delete("/admin/seminars/{seminar_id}")
async def delete_seminar_compat(seminar_id: str, admin: dict = Depends(require_admin)):
    await db.seminars.delete_one({"id": seminar_id})
    return {"message": "Deleted"}

# ============= COMMUNITY ADMIN =============

@api_router.post("/community/posts")
async def create_post(title: str = Form(...), content: str = Form(...), category: str = Form("tartışma")):
    post = {
        "id": str(uuid.uuid4()), "author_email": "anonymous@proptech.com",
        "title": title, "content": content, "category": category,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.community_posts.insert_one(post)
    post.pop("_id", None)
    return post

@api_router.get("/community/posts")
async def get_posts(category: Optional[str] = None):
    query = {} if not category else {"category": category}
    return await db.community_posts.find(query, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)

@api_router.delete("/admin/community/posts/{post_id}")
async def delete_post(post_id: str, admin: dict = Depends(require_admin)):
    result = await db.community_posts.delete_one({"id": post_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted"}

# ============= LAND OPPORTUNITIES ADMIN =============

@api_router.post("/admin/opportunities")
async def create_opportunity(
    admin: dict = Depends(require_admin),
    location_text: str = Form(...), parcel_size_sqm: float = Form(0), zoning_type: str = Form(""),
    investment_potential: str = Form("orta"), risk_score: int = Form(5),
    development_potential: str = Form(""), price_per_sqm: float = Form(0),
    location_lat: float = Form(41.0082), location_lng: float = Form(28.9784),
):
    opp = {
        "id": str(uuid.uuid4()), "location": location_text, "parcel_size_sqm": parcel_size_sqm,
        "zoning_type": zoning_type, "investment_potential": investment_potential,
        "risk_score": risk_score, "development_potential": development_potential,
        "price_per_sqm": price_per_sqm if price_per_sqm > 0 else None,
        "location_coords": {"lat": location_lat, "lng": location_lng},
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.land_opportunities.insert_one(opp)
    opp.pop("_id", None)
    return opp

@api_router.get("/opportunities")
async def get_opportunities():
    return await db.land_opportunities.find({}, {"_id": 0}).limit(500).to_list(500)

@api_router.delete("/admin/opportunities/{opp_id}")
async def delete_opportunity(opp_id: str, admin: dict = Depends(require_admin)):
    result = await db.land_opportunities.delete_one({"id": opp_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted"}

# ============= MARKET DATA ADMIN =============

@api_router.post("/admin/market-data")
async def create_market_data(
    admin: dict = Depends(require_admin),
    neighborhood: str = Form(...), city: str = Form(...), district: str = Form(""),
    avg_price_per_sqm: float = Form(0), price_change_percentage: float = Form(0),
    data_date: str = Form(""),
):
    data_item = {
        "id": str(uuid.uuid4()), "neighborhood": neighborhood, "city": city,
        "district": district, "avg_price_per_sqm": avg_price_per_sqm,
        "price_change_percentage": price_change_percentage, "data_date": data_date,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.market_data.insert_one(data_item)
    data_item.pop("_id", None)
    return data_item

@api_router.get("/market-data")
async def get_market_data(city: Optional[str] = None):
    query = {} if not city else {"city": {"$regex": city, "$options": "i"}}
    return await db.market_data.find(query, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)

@api_router.delete("/admin/market-data/{data_id}")
async def delete_market_data(data_id: str, admin: dict = Depends(require_admin)):
    result = await db.market_data.delete_one({"id": data_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"message": "Deleted"}

# ============= YATIRIM FONU =============

class YatirimFonuBasvuruCreate(BaseModel):
    ad_soyad: str
    telefon: str
    email: str
    sehir: str = ""
    meslek: str = ""
    yatirim_butcesi: str
    ilgi_duyulan_bolge: str
    yatirim_suresi: str
    aciklama: str = ""
    genel_bilgilendirme_onay: bool = False
    iletisim_onay: bool = False

class BeklemeListesiCreate(BaseModel):
    ad_soyad: str
    telefon_veya_email: str

@api_router.post("/yatirim-fonu/basvuru")
async def create_yatirim_fonu_basvuru(data: YatirimFonuBasvuruCreate):
    basvuru = {
        "id": str(uuid.uuid4()),
        **data.dict(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "pending",
    }
    await db.yatirim_fonu_basvurulari.insert_one(basvuru)
    basvuru.pop("_id", None)
    return {"message": "Başvurunuz alınmıştır.", "id": basvuru["id"]}

@api_router.post("/yatirim-fonu/bekleme-listesi")
async def join_bekleme_listesi(data: BeklemeListesiCreate):
    kayit = {
        "id": str(uuid.uuid4()),
        **data.dict(),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.yatirim_fonu_bekleme.insert_one(kayit)
    kayit.pop("_id", None)
    return {"message": "Bekleme listesine eklendiniz.", "id": kayit["id"]}

@api_router.get("/admin/yatirim-fonu/basvurular")
async def get_yatirim_fonu_basvurulari(admin: dict = Depends(require_admin)):
    return await db.yatirim_fonu_basvurulari.find({}, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)

@api_router.get("/admin/yatirim-fonu/bekleme-listesi")
async def get_bekleme_listesi_admin(admin: dict = Depends(require_admin)):
    return await db.yatirim_fonu_bekleme.find({}, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)

# ============= ADMIN STATS =============

@api_router.get("/admin/stats")
async def get_admin_stats(admin: dict = Depends(require_admin)):
    return {
        "projects": await db.projects.count_documents({}),
        "land_parcels": await db.land_parcels.count_documents({}),
        "mega_projects": await db.mega_projects.count_documents({}),
        "courses": await db.courses.count_documents({}),
        "seminars": await db.seminars.count_documents({}),
        "community_posts": await db.community_posts.count_documents({}),
        "opportunities": await db.land_opportunities.count_documents({}),
        "market_data": await db.market_data.count_documents({}),
        "app_users": await db.app_users.count_documents({}),
    }

# ==================== AI AGENT ====================

AGENT_SYSTEM_PROMPT = """Sen bir e-Konut Veri Yönetim Asistanısın. Kullanıcı sana Türkçe komutlar gönderecek ve sen bu komutları anlayıp yapılacak veritabanı işlemlerini JSON formatında döndüreceksin.

Desteklenen işlemler:
- create: Yeni proje oluştur
- bulk_create: Birden fazla proje oluştur (data.projects array içinde)
- update: Mevcut projeyi güncelle
- delete: Projeyi sil
- query: Projeleri listele/sorgula

Proje alanları: project_name (zorunlu), city (İl), district (İlçe), neighborhood (Mahalle), project_type (TOKİ/Emlak Konut/Özel Proje), total_housing, commercial_count, school_count, mosque_count, social_facility_count, progress_percentage (0-100), start_date, planned_end_date, description

SADECE geçerli JSON döndür:
{"message":"Türkçe açıklama","actions":[{"type":"create|bulk_create|update|delete|query","data":{...},"search_name":"güncelle/sil için proje adı","filters":{"city":"...","district":"..."}}]}

bulk_create için: "data":{"projects":[{...},{...}]}
update/delete için: "search_name" zorunlu"""

class AgentMessage(BaseModel):
    message: str
    session_id: str = "default"

@api_router.post("/admin/agent/chat")
async def agent_chat(body: AgentMessage, admin: dict = Depends(require_admin)):
    llm_key = os.environ.get('EMERGENT_LLM_KEY')
    if not llm_key:
        raise HTTPException(status_code=500, detail="LLM key yapılandırılmamış")

    projects_cursor = db.projects.find({}, {"_id": 0, "id": 1, "project_name": 1, "city": 1, "district": 1, "progress_percentage": 1}).limit(200)
    existing_projects = await projects_cursor.to_list(length=200)
    context_msg = f"\nMevcut projeler: {json.dumps([{'id':p['id'],'ad':p.get('project_name'),'il':p.get('city'),'ilce':p.get('district')} for p in existing_projects], ensure_ascii=False)}"

    chat = LlmChat(api_key=llm_key, session_id=f"agent_{body.session_id}", system_message=AGENT_SYSTEM_PROMPT).with_model("anthropic", "claude-sonnet-4-5-20250929")
    response_text = await chat.send_message(LlmUserMessage(text=body.message + context_msg))

    agent_response = None  # initialize before try to prevent undefined-variable risk
    try:
        raw = response_text.strip()
        if "```json" in raw: raw = raw.split("```json")[1].split("```")[0].strip()
        elif "```" in raw: raw = raw.split("```")[1].split("```")[0].strip()
        agent_response = json.loads(raw)
    except Exception as e:
        return {"message": f"Yanıt işlenemedi: {str(e)}", "actions": [], "results": []}

    if not agent_response:
        return {"message": "İşlem tamamlandı.", "results": []}

    results = []
    for action in agent_response.get("actions", []):
        atype = action.get("type")
        try:
            if atype in ("create", "bulk_create"):
                items = action.get("data", {}).get("projects", [action.get("data", {})]) if atype == "bulk_create" else [action.get("data", {})]
                created = []
                for d in items:
                    p = {"id": str(uuid.uuid4()), "project_name": d.get("project_name",""), "city": d.get("city",""), "district": d.get("district",""), "neighborhood": d.get("neighborhood",""), "description": d.get("description",""), "project_type": d.get("project_type","TOKİ"), "total_housing": int(d.get("total_housing",0)), "commercial_count": int(d.get("commercial_count",0)), "school_count": int(d.get("school_count",0)), "mosque_count": int(d.get("mosque_count",0)), "social_facility_count": int(d.get("social_facility_count",0)), "progress_percentage": int(d.get("progress_percentage",0)), "start_date": d.get("start_date",""), "planned_end_date": d.get("planned_end_date",""), "created_at": datetime.now(timezone.utc).isoformat()}
                    await db.projects.insert_one(p)
                    created.append(p["project_name"])
                results.append({"type": atype, "status": "ok", "count": len(created), "names": created})
            elif atype == "update":
                sname = action.get("search_name","")
                existing = await db.projects.find_one({"project_name": {"$regex": sname, "$options": "i"}}, {"_id": 0})
                if existing:
                    await db.projects.update_one({"id": existing["id"]}, {"$set": {k:v for k,v in action.get("data",{}).items() if v is not None and k!="id"}})
                    results.append({"type": "update", "status": "ok", "name": existing["project_name"]})
                else:
                    results.append({"type": "update", "status": "not_found", "search": sname})
            elif atype == "delete":
                sname = action.get("search_name","")
                existing = await db.projects.find_one({"project_name": {"$regex": sname, "$options": "i"}}, {"_id": 0})
                if existing:
                    await db.projects.delete_one({"id": existing["id"]})
                    results.append({"type": "delete", "status": "ok", "name": existing["project_name"]})
                else:
                    results.append({"type": "delete", "status": "not_found", "search": sname})
            elif atype == "query":
                f = action.get("filters", {})
                q = {k: {"$regex": v, "$options": "i"} for k, v in [("city", f.get("city")), ("district", f.get("district"))] if v}
                if f.get("project_type"): q["project_type"] = f["project_type"]
                found = await db.projects.find(q, {"_id": 0, "id": 1, "project_name": 1, "city": 1, "district": 1, "progress_percentage": 1, "total_housing": 1}).to_list(50)
                results.append({"type": "query", "status": "ok", "count": len(found), "projects": found})
        except Exception as e:
            results.append({"type": atype, "status": "error", "error": str(e)})

    return {"message": agent_response.get("message","İşlem tamamlandı."), "results": results}

# ---- Agent ZIP Upload ----
@api_router.post("/admin/agent/upload-zip")
async def agent_upload_zip(
    file: UploadFile = File(...),
    project_id: str = Form(None),     # Optional – if omitted, a new project is created
    message: str = Form(""),          # User's description text (used when creating new project)
    admin: dict = Depends(require_admin)
):
    """Accept a ZIP/RAR (or single KML/KMZ/GeoJSON/image) and process all contents.
    If no project_id is provided, a new project is created from message + DOCX content.
    """
    import zipfile
    import io as _io
    import rarfile
    from PIL import Image as PilImage

    raw = await file.read()
    fname = file.filename or "upload"
    ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""

    GEO_EXTS  = {"kml", "kmz", "geojson", "json"}
    IMG_EXTS  = {"jpg", "jpeg", "png", "webp"}
    DOC_EXTS  = {"docx"}
    MIME_TYPES = {"kml":"application/vnd.google-earth.kml+xml","kmz":"application/vnd.google-earth.kmz","geojson":"application/geo+json","json":"application/json"}

    # --- Collect (basename, ext, bytes) from archive or single file ---
    all_files = []

    def _add(name, data):
        bname = name.replace("\\", "/").split("/")[-1]
        if not bname or bname.startswith(".") or bname.lower() == "thumbs.db":
            return
        fext = bname.rsplit(".", 1)[-1].lower() if "." in bname else ""
        if fext in GEO_EXTS | IMG_EXTS | DOC_EXTS:
            all_files.append((bname, fext, data))

    if ext == "zip":
        with zipfile.ZipFile(_io.BytesIO(raw)) as z:
            for name in z.namelist():
                if not name.startswith("__MACOSX") and not name.endswith("/"):
                    _add(name, z.read(name))
    elif ext == "rar":
        import tempfile
        import os as _os
        with tempfile.NamedTemporaryFile(suffix=".rar", delete=False) as tmp:
            tmp.write(raw)
            tmp_path = tmp.name
        try:
            with rarfile.RarFile(tmp_path) as rf:
                for info in rf.infolist():
                    if not info.is_dir():
                        _add(info.filename, rf.read(info.filename))
        finally:
            _os.unlink(tmp_path)
    elif ext in GEO_EXTS | IMG_EXTS:
        all_files.append((fname, ext, raw))
    else:
        raise HTTPException(status_code=400, detail="ZIP, RAR, KML, KMZ, GeoJSON veya görsel dosyası yükleyin")

    if not all_files:
        raise HTTPException(status_code=400, detail="Desteklenen dosya bulunamadı (KML/KMZ/GeoJSON/JPG/PNG/DOCX)")

    # --- Extract DOCX text early (needed for new-project creation) ---
    docx_full_text = ""
    for (orig_name, fext, data) in all_files:
        if fext == "docx" and not docx_full_text:
            try:
                from docx import Document as DocxDoc
                doc = DocxDoc(_io.BytesIO(data))
                lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
                docx_full_text = "\n".join(lines[:40])
            except Exception:
                pass

    # --- If no project_id provided → create a new project via LLM ---
    created_project = None
    if not project_id:
        llm_key = os.environ.get("EMERGENT_LLM_KEY")
        project_info = {}

        combined_input = ""
        if message.strip():
            combined_input += f"Kullanıcı açıklaması:\n{message.strip()}\n\n"
        if docx_full_text:
            combined_input += f"DOCX dosyasından elde edilen metin:\n{docx_full_text[:2000]}\n\n"
        if not combined_input:
            combined_input = f"ZIP dosya adı: {fname}"

        if llm_key:
            try:
                extraction_prompt = f"""Aşağıdaki bilgilerden e-Konut projesi verilerini çıkar ve SADECE geçerli JSON döndür:

{combined_input}

Çıkarılacak JSON şeması (bilinmeyenler için boş bırak, sayılar için 0 kullan):
{{
  "project_name": "Zorunlu - proje adı",
  "city": "İl adı",
  "district": "İlçe adı",
  "neighborhood": "Mahalle adı",
  "description": "Proje açıklaması",
  "project_type": "TOKİ veya Emlak Konut veya Özel Proje",
  "total_housing": 0,
  "progress_percentage": 0,
  "start_date": "YYYY-MM-DD veya boş",
  "planned_end_date": "YYYY-MM-DD veya boş"
}}

Önemli: project_name zorunludur. Bulunamazsa dosya adından türet: {fname.rsplit('.', 1)[0].replace('-', ' ').replace('_', ' ').title()}"""

                from emergentintegrations.llm.chat import LlmChat, UserMessage as LlmUserMessage
                chat = LlmChat(
                    api_key=llm_key,
                    session_id=f"zip_extract_{uuid.uuid4()}",
                    system_message="Sen bir veri çıkarma asistanısın. Sadece geçerli JSON döndür, başka hiçbir şey yazma."
                ).with_model("anthropic", "claude-sonnet-4-5-20250929")
                raw_resp = await chat.send_message(LlmUserMessage(text=extraction_prompt))
                raw_resp = raw_resp.strip()
                if "```json" in raw_resp:
                    raw_resp = raw_resp.split("```json")[1].split("```")[0].strip()
                elif "```" in raw_resp:
                    raw_resp = raw_resp.split("```")[1].split("```")[0].strip()
                project_info = json.loads(raw_resp)
            except Exception:
                project_info = {}

        # Fallback name from filename
        if not project_info.get("project_name"):
            project_info["project_name"] = fname.rsplit(".", 1)[0].replace("-", " ").replace("_", " ").title()

        new_project = {
            "id": str(uuid.uuid4()),
            "project_name": project_info.get("project_name", "Yeni Proje"),
            "city":         project_info.get("city", ""),
            "district":     project_info.get("district", ""),
            "neighborhood": project_info.get("neighborhood", ""),
            "description":  project_info.get("description", docx_full_text[:500] if docx_full_text else ""),
            "project_type": project_info.get("project_type", "TOKİ"),
            "total_housing":         int(project_info.get("total_housing", 0) or 0),
            "commercial_count":      0,
            "school_count":          0,
            "mosque_count":          0,
            "social_facility_count": 0,
            "progress_percentage":   int(project_info.get("progress_percentage", 0) or 0),
            "start_date":       project_info.get("start_date", ""),
            "planned_end_date": project_info.get("planned_end_date", ""),
            "youtube_videos":   [],
            "created_at": datetime.now(timezone.utc).isoformat(),
        }
        await db.projects.insert_one(new_project)
        new_project.pop("_id", None)
        project_id = new_project["id"]
        created_project = new_project
    else:
        # Verify existing project
        existing = await db.projects.find_one({"id": project_id}, {"_id": 0})
        if not existing:
            raise HTTPException(status_code=404, detail="Proje bulunamadı")

    # --- Process all files ---
    added_layers  = []
    added_images  = []
    description_text = docx_full_text or None
    project_center   = None

    def _compress_image(data: bytes) -> bytes:
        img = PilImage.open(_io.BytesIO(data))
        img.thumbnail((1280, 960), PilImage.LANCZOS)
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")
        buf = _io.BytesIO()
        img.save(buf, "JPEG", quality=82, optimize=True)
        return buf.getvalue()

    for (orig_name, fext, data) in all_files:
        # Geo files → map layers
        if fext in GEO_EXTS:
            storage_path = f"{APP_NAME}/map-layers/{project_id}/{uuid.uuid4()}.{fext}"
            content_type = MIME_TYPES.get(fext, "application/octet-stream")
            put_object(storage_path, data, content_type)
            layer = {
                "id": str(uuid.uuid4()),
                "project_id": project_id,
                "storage_path": storage_path,
                "original_filename": orig_name,
                "content_type": content_type,
                "file_type": fext.upper(),
                "size": len(data),
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.project_map_layers.insert_one(layer)
            added_layers.append(orig_name)
            if project_center is None:
                try:
                    if fext in ("kml", "kmz"):
                        project_center = extract_centroid_from_kml(fext, data)
                    elif fext in ("geojson", "json"):
                        project_center = extract_centroid_from_geojson(data)
                except Exception:
                    pass

        # Images → project media (auto-resized)
        elif fext in IMG_EXTS:
            try:
                compressed = _compress_image(data)
            except Exception:
                compressed = data
            storage_path = f"{APP_NAME}/media/{project_id}/{uuid.uuid4()}.jpg"
            put_object(storage_path, compressed, "image/jpeg")
            media_doc = {
                "id": str(uuid.uuid4()),
                "project_id": project_id,
                "storage_path": storage_path,
                "original_filename": orig_name,
                "media_type": "IMAGE",
                "category": "Görsel",
                "title": orig_name,
                "content_type": "image/jpeg",
                "size": len(compressed),
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.project_media.insert_one(media_doc)
            added_images.append(orig_name)

    # Save location & description to project
    update_fields = {}
    if project_center:
        update_fields["location"] = project_center
    if description_text and not created_project:
        # For new projects description already set; for existing only fill if empty
        existing_desc = (await db.projects.find_one({"id": project_id}, {"_id": 0, "description": 1}) or {}).get("description", "")
        if not existing_desc:
            update_fields["description"] = description_text[:500]
    if update_fields:
        await db.projects.update_one({"id": project_id}, {"$set": update_fields})

    project_doc = await db.projects.find_one({"id": project_id}, {"_id": 0, "project_name": 1})
    project_name = project_doc.get("project_name", "") if project_doc else (created_project or {}).get("project_name", "")

    response = {
        "added_layers":          added_layers,
        "added_images":          added_images,
        "layers_count":          len(added_layers),
        "images_count":          len(added_images),
        "project_name":          project_name,
        "location_updated":      project_center is not None,
        "description_extracted": description_text is not None,
    }
    if created_project:
        response["new_project"] = {
            "id":   created_project["id"],
            "name": created_project["project_name"],
            "city": created_project.get("city", ""),
            "district": created_project.get("district", ""),
        }
    return response

# ============= SHARED FACILITIES =============

@api_router.get("/shared-facilities")
async def get_shared_facilities():
    docs = await db.shared_facilities.find({}, {"_id": 0}).to_list(500)
    return docs

@api_router.post("/admin/shared-facilities")
async def create_shared_facility(body: dict, admin: dict = Depends(require_admin)):
    facility = {
        "id": str(uuid.uuid4()),
        "name": body.get("name", ""),
        "type": body.get("type", "diger"),
        "lat": float(body.get("lat", 0)),
        "lng": float(body.get("lng", 0)),
        "description": body.get("description", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.shared_facilities.insert_one(facility)
    facility.pop("_id", None)
    return facility

@api_router.delete("/admin/shared-facilities/{facility_id}")
async def delete_shared_facility(facility_id: str, admin: dict = Depends(require_admin)):
    result = await db.shared_facilities.delete_one({"id": facility_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Tesis bulunamadı")
    return {"message": "Silindi"}

@api_router.put("/admin/shared-facilities/{facility_id}")
async def update_shared_facility(facility_id: str, body: dict, admin: dict = Depends(require_admin)):
    update = {k: v for k, v in body.items() if k in ("name", "type", "lat", "lng", "description")}
    if "lat" in update: update["lat"] = float(update["lat"])
    if "lng" in update: update["lng"] = float(update["lng"])
    await db.shared_facilities.update_one({"id": facility_id}, {"$set": update})
    return {"message": "Güncellendi"}




# ============= MRXAKADEMİ KAPSAMLI YÖNETİM =============

@api_router.get("/admin/academy-stats")
async def admin_academy_stats(admin: dict = Depends(require_admin)):
    courses = await db.courses.count_documents({"status": "active"})
    seminars = await db.seminars.count_documents({})
    students = await db.app_users.count_documents({"role": "user"})
    streams = await db.live_streams.count_documents({})
    supervision = await db.supervision_events.count_documents({})
    exams = await db.course_exams.count_documents({})
    payments = await db.user_payments.count_documents({})
    contracts = await db.user_contracts.count_documents({})
    files = await db.user_files.count_documents({})
    exam_attempts = await db.user_exam_attempts.count_documents({})
    passed = await db.user_exam_attempts.count_documents({"passed": True})
    return {
        "courses": courses, "seminars": seminars, "students": students,
        "streams": streams, "supervision": supervision, "exams": exams,
        "payments": payments, "contracts": contracts, "files": files,
        "exam_attempts": exam_attempts, "passed_exams": passed,
    }

@api_router.get("/admin/students")
async def admin_get_students(admin: dict = Depends(require_admin)):
    users = await db.app_users.find({"role": "user"}, {"_id": 0, "password": 0, "hashed_password": 0}).to_list(1000)
    result = []
    for user in users:
        uid = user.get("user_id")
        prog_records = await db.user_progress.find({"user_id": uid}, {"_id": 0}).to_list(100)
        attempts = await db.user_exam_attempts.find({"user_id": uid}, {"_id": 0}).to_list(100)
        completed_lessons = sum(len(p.get("completed_lessons", [])) for p in prog_records)
        last_activity = max([p.get("updated_at", "") for p in prog_records], default=None)
        best_score = max([a.get("score", 0) for a in attempts], default=0)
        result.append({
            **{k: v for k, v in user.items() if k not in ("_id",)},
            "completed_lessons": completed_lessons,
            "exam_attempts": len(attempts),
            "best_score": best_score,
            "courses_enrolled": len(prog_records),
            "last_activity": last_activity,
        })
    return result

@api_router.get("/admin/students/{user_id}")
async def admin_get_student_detail(user_id: str, admin: dict = Depends(require_admin)):
    user = await db.app_users.find_one({"user_id": user_id}, {"_id": 0, "password": 0, "hashed_password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    courses = await db.courses.find({"status": "active"}, {"_id": 0, "modules": 1, "title": 1, "id": 1, "cover_image": 1}).to_list(100)
    progress_list = []
    for course in courses:
        prog = await db.user_progress.find_one({"user_id": user_id, "course_id": course.get("id")}, {"_id": 0})
        total = sum(len(m.get("lessons", [])) for m in course.get("modules", []))
        completed = len(prog.get("completed_lessons", [])) if prog else 0
        pct = round(completed / total * 100) if total > 0 else 0
        progress_list.append({
            "course_id": course.get("id"), "title": course.get("title"),
            "cover_image": course.get("cover_image", ""), "progress_pct": pct,
            "completed": completed, "total": total,
            "last_lesson": prog.get("last_lesson_id") if prog else None
        })
    attempts = await db.user_exam_attempts.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    # Enrich attempts with exam titles
    for a in attempts:
        exam = await db.course_exams.find_one({"id": a.get("exam_id")}, {"_id": 0, "title": 1})
        a["exam_title"] = exam.get("title", "Sınav") if exam else "Sınav"
    files = await db.user_files.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    payments = await db.user_payments.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    contracts = await db.user_contracts.find({"user_id": user_id}, {"_id": 0}).to_list(100)
    return {"user": user, "progress": progress_list, "exam_attempts": attempts, "files": files, "payments": payments, "contracts": contracts}

@api_router.get("/admin/payments")
async def admin_get_all_payments(admin: dict = Depends(require_admin)):
    payments = await db.user_payments.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return payments

@api_router.post("/admin/payments")
async def admin_create_payment(body: dict, admin: dict = Depends(require_admin)):
    payment = {
        "id": str(uuid.uuid4()), "user_id": body.get("user_id", ""), "course_name": body.get("course_name", ""),
        "amount": body.get("amount", ""), "status": body.get("status", "pending"),
        "notes": body.get("notes", ""), "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_payments.insert_one(payment)
    payment.pop("_id", None)
    return payment

@api_router.put("/admin/payments/{payment_id}")
async def admin_update_payment(payment_id: str, body: dict, admin: dict = Depends(require_admin)):
    update = {k: v for k, v in body.items() if k in ("course_name", "amount", "status", "notes")}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.user_payments.update_one({"id": payment_id}, {"$set": update})
    return {"message": "Güncellendi"}

@api_router.delete("/admin/payments/{payment_id}")
async def admin_delete_payment(payment_id: str, admin: dict = Depends(require_admin)):
    await db.user_payments.delete_one({"id": payment_id})
    return {"message": "Silindi"}

@api_router.get("/admin/contracts")
async def admin_get_all_contracts(admin: dict = Depends(require_admin)):
    contracts = await db.user_contracts.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return contracts

@api_router.post("/admin/contracts")
async def admin_create_contract(body: dict, admin: dict = Depends(require_admin)):
    contract = {
        "id": str(uuid.uuid4()), "user_id": body.get("user_id", ""), "contract_name": body.get("contract_name", ""),
        "status": body.get("status", "pending"), "notes": body.get("notes", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_contracts.insert_one(contract)
    contract.pop("_id", None)
    return contract

@api_router.put("/admin/contracts/{contract_id}")
async def admin_update_contract(contract_id: str, body: dict, admin: dict = Depends(require_admin)):
    update = {k: v for k, v in body.items() if k in ("contract_name", "status", "notes")}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.user_contracts.update_one({"id": contract_id}, {"$set": update})
    return {"message": "Güncellendi"}

@api_router.delete("/admin/contracts/{contract_id}")
async def admin_delete_contract(contract_id: str, admin: dict = Depends(require_admin)):
    await db.user_contracts.delete_one({"id": contract_id})
    return {"message": "Silindi"}

@api_router.get("/admin/all-files")
async def admin_get_all_files(admin: dict = Depends(require_admin)):
    files = await db.user_files.find({}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return files

@api_router.post("/admin/files/{user_id}")
async def admin_add_file(user_id: str, body: dict, admin: dict = Depends(require_admin)):
    f = {
        "id": str(uuid.uuid4()), "user_id": user_id, "file_name": body.get("file_name", ""),
        "file_url": body.get("file_url", ""), "file_type": body.get("file_type", "document"),
        "status": "active", "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.user_files.insert_one(f)
    f.pop("_id", None)
    return f

@api_router.delete("/admin/files/{file_id}")
async def admin_delete_file_admin(file_id: str, admin: dict = Depends(require_admin)):
    await db.user_files.delete_one({"id": file_id})
    return {"message": "Silindi"}

# ============= AI SINAV ÇIKARICI (PDF → Sorular) =============

@api_router.post("/admin/exams/extract-from-pdf")
async def extract_exam_from_pdf(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    llm_key = os.environ.get("EMERGENT_LLM_KEY", "")
    if not llm_key:
        raise HTTPException(status_code=500, detail="LLM key bulunamadı")
    if not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Sadece PDF dosyası yükleyin")

    # Save to temp file
    contents = await file.read()
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        tmp.write(contents)
        tmp_path = tmp.name

    try:
        from emergentintegrations.llm.chat import FileContentWithMimeType
        chat = LlmChat(
            api_key=llm_key,
            session_id=f"exam_extract_{uuid.uuid4()}",
            system_message="""Sen bir eğitim içeriği asistanısın. 
PDF dosyasındaki sınav sorularını çıkarıp JSON formatında döndürüyorsun.
SADECE geçerli JSON döndür, başka hiçbir şey yazma. Türkçe içeriği koru."""
        ).with_model("gemini", "gemini-2.5-flash")

        pdf_file = FileContentWithMimeType(file_path=tmp_path, mime_type="application/pdf")

        prompt = """Bu PDF dosyasındaki sınav içeriğini analiz et ve aşağıdaki JSON formatında döndür:

{
  "title": "Sınav başlığı (yoksa PDF adından türet)",
  "pass_score": 70,
  "duration_minutes": 30,
  "questions": [
    {
      "id": "q1",
      "text": "Soru metni",
      "options": ["Şık A", "Şık B", "Şık C", "Şık D"],
      "correct_answer": "Doğru şıkkın tam metni"
    }
  ]
}

Önemli kurallar:
- Tüm sorular için tam 4 şık çıkar
- correct_answer, options listesindeki tam metinle eşleşmeli
- Şıklar A/B/C/D veya 1/2/3/4 şeklindeyse saf metinleri al (harf/rakam prefix olmadan)
- En az 1 soru çıkar, maksimum tüm soruları çıkar
- Sadece JSON döndür, açıklama veya markdown ekleme"""

        user_message = LlmUserMessage(text=prompt, file_contents=[pdf_file])
        response = await chat.send_message(user_message)

        # Parse JSON — remove any markdown wrapper if present
        raw = response.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()

        data = json.loads(raw)

        # Ensure IDs are set
        for i, q in enumerate(data.get("questions", []), 1):
            q["id"] = f"q{i}"
            # Ensure exactly 4 options
            if len(q.get("options", [])) < 4:
                q["options"] = q["options"] + [""] * (4 - len(q["options"]))

        return data

    except json.JSONDecodeError as e:
        logger.error(f"PDF JSON parse error: {e} | raw: {response[:200]}")
        raise HTTPException(status_code=422, detail="AI yanıtı JSON olarak ayrıştırılamadı")
    except Exception as e:
        logger.error(f"PDF extraction error: {e}")
        raise HTTPException(status_code=500, detail=f"PDF işlenirken hata oluştu: {str(e)[:100]}")
    finally:
        import os as _os
        try:
            _os.unlink(tmp_path)
        except Exception:
            pass

# ============= ADMIN PANEL YÖNETİMİ (Canlı Yayın, Süpervizyon, Sınavlar) =============

@api_router.get("/admin/live-streams")
async def admin_get_live_streams(admin: dict = Depends(require_admin)):
    streams = await db.live_streams.find({}, {"_id": 0}).sort("date", -1).to_list(200)
    return streams

@api_router.post("/admin/live-streams")
async def admin_create_live_stream(body: dict, admin: dict = Depends(require_admin)):
    stream = {
        "id": str(uuid.uuid4()),
        "title": body.get("title", ""),
        "date": body.get("date", ""),
        "status": body.get("status", "upcoming"),
        "platform": body.get("platform", "Zoom"),
        "join_url": body.get("join_url", ""),
        "thumbnail": body.get("thumbnail", ""),
        "description": body.get("description", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.live_streams.insert_one(stream)
    stream.pop("_id", None)
    return stream

@api_router.put("/admin/live-streams/{stream_id}")
async def admin_update_live_stream(stream_id: str, body: dict, admin: dict = Depends(require_admin)):
    update = {k: v for k, v in body.items() if k in ("title", "date", "status", "platform", "join_url", "thumbnail", "description")}
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.live_streams.update_one({"id": stream_id}, {"$set": update})
    return {"message": "Güncellendi"}

@api_router.delete("/admin/live-streams/{stream_id}")
async def admin_delete_live_stream(stream_id: str, admin: dict = Depends(require_admin)):
    await db.live_streams.delete_one({"id": stream_id})
    return {"message": "Silindi"}

@api_router.get("/admin/supervision")
async def admin_get_supervision(admin: dict = Depends(require_admin)):
    events = await db.supervision_events.find({}, {"_id": 0}).sort("date", 1).to_list(200)
    return events

@api_router.post("/admin/supervision")
async def admin_create_supervision(body: dict, admin: dict = Depends(require_admin)):
    event = {
        "id": str(uuid.uuid4()),
        "title": body.get("title", ""),
        "location": body.get("location", ""),
        "city": body.get("city", ""),
        "date": body.get("date", ""),
        "status": body.get("status", "upcoming"),
        "capacity": int(body.get("capacity", 20)),
        "registered": int(body.get("registered", 0)),
        "description": body.get("description", ""),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.supervision_events.insert_one(event)
    event.pop("_id", None)
    return event

@api_router.put("/admin/supervision/{event_id}")
async def admin_update_supervision(event_id: str, body: dict, admin: dict = Depends(require_admin)):
    update = {k: v for k, v in body.items() if k in ("title", "location", "city", "date", "status", "capacity", "registered", "description")}
    if "capacity" in update: update["capacity"] = int(update["capacity"])
    if "registered" in update: update["registered"] = int(update["registered"])
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.supervision_events.update_one({"id": event_id}, {"$set": update})
    return {"message": "Güncellendi"}

@api_router.delete("/admin/supervision/{event_id}")
async def admin_delete_supervision(event_id: str, admin: dict = Depends(require_admin)):
    await db.supervision_events.delete_one({"id": event_id})
    return {"message": "Silindi"}

@api_router.get("/admin/exams")
async def admin_get_exams(admin: dict = Depends(require_admin)):
    exams = await db.course_exams.find({}, {"_id": 0}).to_list(200)
    return exams

@api_router.post("/admin/exams")
async def admin_create_exam(body: dict, admin: dict = Depends(require_admin)):
    exam = {
        "id": str(uuid.uuid4()),
        "course_id": body.get("course_id", ""),
        "title": body.get("title", ""),
        "pass_score": int(body.get("pass_score", 70)),
        "duration_minutes": int(body.get("duration_minutes", 30)),
        "questions": body.get("questions", []),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.course_exams.insert_one(exam)
    exam.pop("_id", None)
    return exam

@api_router.put("/admin/exams/{exam_id}")
async def admin_update_exam(exam_id: str, body: dict, admin: dict = Depends(require_admin)):
    update = {k: v for k, v in body.items() if k in ("course_id", "title", "pass_score", "duration_minutes", "questions")}
    if "pass_score" in update: update["pass_score"] = int(update["pass_score"])
    if "duration_minutes" in update: update["duration_minutes"] = int(update["duration_minutes"])
    update["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.course_exams.update_one({"id": exam_id}, {"$set": update})
    return {"message": "Güncellendi"}

@api_router.delete("/admin/exams/{exam_id}")
async def admin_delete_exam(exam_id: str, admin: dict = Depends(require_admin)):
    await db.course_exams.delete_one({"id": exam_id})
    return {"message": "Silindi"}

@api_router.get("/admin/panel-stats")
async def admin_panel_stats(admin: dict = Depends(require_admin)):
    streams = await db.live_streams.count_documents({})
    supervision = await db.supervision_events.count_documents({})
    exams = await db.course_exams.count_documents({})
    panel_users = await db.user_progress.distinct("user_id")
    return {"streams": streams, "supervision": supervision, "exams": exams, "active_users": len(panel_users)}

# ============= USER PANEL APIs =============

@api_router.get("/user/profile")
async def get_user_profile(request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    doc = await db.app_users.find_one({"user_id": user["user_id"]}, {"_id": 0, "password": 0, "hashed_password": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Kullanıcı bulunamadı")
    return {
        "user_id": doc.get("user_id"),
        "full_name": doc.get("full_name", ""),
        "email": doc.get("email", ""),
        "phone": doc.get("phone", ""),
        "avatar_color": doc.get("avatar_color", "emerald"),
        "plan": doc.get("plan", "free"),
        "created_at": doc.get("created_at", ""),
        "auth_provider": doc.get("auth_provider", "email"),
    }

@api_router.put("/user/profile")
async def update_user_profile(body: dict, request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    allowed = {}
    if body.get("full_name", "").strip():
        allowed["full_name"] = body["full_name"].strip()
    if "phone" in body:
        allowed["phone"] = body["phone"].strip()
    if "avatar_color" in body:
        allowed["avatar_color"] = body["avatar_color"]
    allowed["updated_at"] = datetime.now(timezone.utc).isoformat()
    await db.app_users.update_one({"user_id": user["user_id"]}, {"$set": allowed})
    doc = await db.app_users.find_one({"user_id": user["user_id"]}, {"_id": 0, "password": 0, "hashed_password": 0})
    return {
        "user_id": doc.get("user_id"),
        "full_name": doc.get("full_name", ""),
        "email": doc.get("email", ""),
        "phone": doc.get("phone", ""),
        "avatar_color": doc.get("avatar_color", "emerald"),
        "plan": doc.get("plan", "free"),
    }

@api_router.get("/packages")
async def get_packages():
    return [
        {
            "id": "free", "name": "Ücretsiz", "price": 0, "period": "",
            "color": "slate", "popular": False,
            "features": ["Ücretsiz seminerlere erişim", "Topluluk forumuna katılım", "Temel içerikler ve makaleler", "Aylık 1 canlı yayın"],
        },
        {
            "id": "basic", "name": "Temel", "price": 499, "period": "ay",
            "color": "emerald", "popular": False,
            "features": ["Tüm aktif kurslar", "Eğitim sınavları ve değerlendirmeler", "Tamamlama sertifikaları", "Dosya indirme erişimi", "E-posta desteği"],
        },
        {
            "id": "pro", "name": "Pro", "price": 999, "period": "ay",
            "color": "blue", "popular": True,
            "features": ["Temel paket + tüm özellikler", "Sınırsız canlı yayın erişimi", "Süpervizyon eğitimleri", "Öncelikli destek hattı", "Özel grup oturumları", "İndirimli seminer biletleri"],
        },
        {
            "id": "corporate", "name": "Kurumsal", "price": 2499, "period": "ay",
            "color": "amber", "popular": False,
            "features": ["Tüm Pro özellikler", "Birebir özel danışmanlık", "Kurumsal grup eğitimleri", "Özelleştirilmiş sözleşme", "7/24 öncelikli destek", "Yıllık yatırım raporu"],
        },
    ]

@api_router.get("/user/progress")
async def get_user_progress(request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    user_id = user.get("user_id")
    courses = await db.courses.find({"status": "active"}, {"_id": 0}).to_list(100)
    result = []
    for course in courses:
        cid = course.get("id")
        prog = await db.user_progress.find_one({"user_id": user_id, "course_id": cid}, {"_id": 0})
        total_lessons = sum(len(m.get("lessons", [])) for m in course.get("modules", []))
        completed = prog.get("completed_lessons", []) if prog else []
        pct = round(len(completed) / total_lessons * 100) if total_lessons > 0 else 0
        result.append({
            "course_id": cid, "title": course.get("title", ""), "cover_image": course.get("cover_image", ""),
            "total_lessons": total_lessons, "completed_lessons": len(completed),
            "progress_pct": pct, "last_lesson_id": prog.get("last_lesson_id") if prog else None,
        })
    return result

@api_router.post("/user/progress/{course_id}/{lesson_id}")
async def update_lesson_progress(course_id: str, lesson_id: str, request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    user_id = user.get("user_id")
    await db.user_progress.update_one(
        {"user_id": user_id, "course_id": course_id},
        {"$addToSet": {"completed_lessons": lesson_id}, "$set": {"last_lesson_id": lesson_id, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return {"message": "İlerleme kaydedildi"}

@api_router.get("/user/files")
async def get_user_files(request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    files = await db.user_files.find({"user_id": user.get("user_id")}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return files

@api_router.delete("/user/files/{file_id}")
async def delete_user_file(file_id: str, request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    await db.user_files.delete_one({"id": file_id, "user_id": user.get("user_id")})
    return {"message": "Dosya silindi"}

@api_router.get("/user/payments")
async def get_user_payments(request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    payments = await db.user_payments.find({"user_id": user.get("user_id")}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return payments

@api_router.get("/user/contracts")
async def get_user_contracts(request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    contracts = await db.user_contracts.find({"user_id": user.get("user_id")}, {"_id": 0}).sort("created_at", -1).to_list(200)
    return contracts

@api_router.get("/user/exams")
async def get_user_exams(request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    user_id = user.get("user_id")
    exams = await db.course_exams.find({}, {"_id": 0}).to_list(100)
    result = []
    for exam in exams:
        attempt = await db.user_exam_attempts.find_one({"user_id": user_id, "exam_id": exam.get("id")}, {"_id": 0})
        result.append({**exam, "attempt": attempt})
    return result

@api_router.post("/user/exams/{exam_id}/submit")
async def submit_exam(exam_id: str, body: dict, request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    exam = await db.course_exams.find_one({"id": exam_id}, {"_id": 0})
    if not exam:
        raise HTTPException(status_code=404, detail="Sınav bulunamadı")
    answers = body.get("answers", {})
    questions = exam.get("questions", [])
    correct = sum(1 for q in questions if answers.get(q["id"]) == q.get("correct_answer"))
    score = round(correct / len(questions) * 100) if questions else 0
    attempt = {
        "id": str(uuid.uuid4()), "user_id": user.get("user_id"), "exam_id": exam_id,
        "score": score, "answers": answers, "completed_at": datetime.now(timezone.utc).isoformat(),
        "passed": score >= exam.get("pass_score", 70)
    }
    await db.user_exam_attempts.insert_one(attempt)
    attempt.pop("_id", None)
    return attempt

@api_router.get("/user/certificates")
async def get_user_certificates(request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    user_id = user.get("user_id")
    courses = await db.courses.find({"status": "active"}, {"_id": 0}).to_list(100)
    result = []
    for course in courses:
        cid = course.get("id")
        prog = await db.user_progress.find_one({"user_id": user_id, "course_id": cid}, {"_id": 0})
        total = sum(len(m.get("lessons", [])) for m in course.get("modules", []))
        completed = len(prog.get("completed_lessons", [])) if prog else 0
        pct = round(completed / total * 100) if total > 0 else 0
        result.append({
            "course_id": cid, "title": course.get("title", ""), "cover_image": course.get("cover_image", ""),
            "progress_pct": pct, "eligible": pct >= 90,
            "certificate_url": f"/certificates/{user_id}/{cid}" if pct >= 90 else None
        })
    return result

@api_router.get("/user/transcript")
async def get_user_transcript(request: Request):
    user = await get_session_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Giriş yapmanız gerekiyor")
    user_id = user.get("user_id")
    courses = await db.courses.find({"status": "active"}, {"_id": 0}).to_list(100)
    result = []
    for course in courses:
        cid = course.get("id")
        prog = await db.user_progress.find_one({"user_id": user_id, "course_id": cid}, {"_id": 0})
        total = sum(len(m.get("lessons", [])) for m in course.get("modules", []))
        completed = len(prog.get("completed_lessons", [])) if prog else 0
        pct = round(completed / total * 100) if total > 0 else 0
        if pct >= 90:
            attempt = await db.user_exam_attempts.find_one({"user_id": user_id}, {"_id": 0}, sort=[("completed_at", -1)])
            result.append({
                "course_id": cid, "title": course.get("title", ""), "progress_pct": pct,
                "score": attempt.get("score", 0) if attempt else 0,
                "completed_at": prog.get("updated_at") if prog else None
            })
    return result

@api_router.get("/live-streams")
async def get_live_streams():
    streams = await db.live_streams.find({}, {"_id": 0}).sort("date", -1).to_list(100)
    if not streams:
        now = datetime.now(timezone.utc)
        streams = [
            {"id": "ls1", "title": "Arsa Yatırımı Temelleri - Canlı Ders", "date": (now + timedelta(days=2)).isoformat(), "status": "upcoming", "platform": "Zoom", "join_url": "#", "thumbnail": ""},
            {"id": "ls2", "title": "Tapu ve Kadastro Hukuku", "date": (now + timedelta(days=7)).isoformat(), "status": "upcoming", "platform": "Zoom", "join_url": "#", "thumbnail": ""},
            {"id": "ls3", "title": "Arsa Değerleme Metodolojisi", "date": now.isoformat(), "status": "live", "platform": "Zoom", "join_url": "#", "thumbnail": ""},
            {"id": "ls4", "title": "İmar Planı Okuma Teknikleri", "date": (now - timedelta(days=5)).isoformat(), "status": "ended", "platform": "Zoom", "join_url": "#", "thumbnail": ""},
            {"id": "ls5", "title": "Yatırım Analizi Workshop", "date": (now - timedelta(days=12)).isoformat(), "status": "ended", "platform": "Zoom", "join_url": "#", "thumbnail": ""},
        ]
    return streams

@api_router.get("/supervision/events")
async def get_supervision_events():
    events = await db.supervision_events.find({}, {"_id": 0}).sort("date", 1).to_list(100)
    if not events:
        now = datetime.now(timezone.utc)
        events = [
            {"id": "sv1", "title": "İstanbul Arsa Analiz Süpervizyon", "location": "İstanbul - Kadıköy Ofis", "city": "İstanbul", "date": (now + timedelta(days=3)).isoformat(), "status": "upcoming", "capacity": 15, "registered": 8},
            {"id": "sv2", "title": "Ankara Mega Proje İnceleme", "location": "Ankara - Çankaya Merkez", "city": "Ankara", "date": (now + timedelta(days=10)).isoformat(), "status": "upcoming", "capacity": 20, "registered": 12},
            {"id": "sv3", "title": "İzmir Kıyı Arsa Gezisi", "location": "İzmir - Alsancak", "city": "İzmir", "date": (now + timedelta(days=18)).isoformat(), "status": "upcoming", "capacity": 10, "registered": 6},
            {"id": "sv4", "title": "Bursa OSB Yatırım Turu", "location": "Bursa - Nilüfer", "city": "Bursa", "date": (now - timedelta(days=7)).isoformat(), "status": "ended", "capacity": 12, "registered": 12},
        ]
    return events

# Exam seed helper (called once on startup)
async def seed_exams():
    count = await db.course_exams.count_documents({})
    if count == 0:
        courses = await db.courses.find({"status": "active"}, {"_id": 0}).to_list(10)
        for course in courses[:3]:
            exam = {
                "id": str(uuid.uuid4()), "course_id": course.get("id"),
                "title": f"{course.get('title', 'Kurs')} - Değerlendirme Sınavı",
                "pass_score": 70, "duration_minutes": 30,
                "questions": [
                    {"id": "q1", "text": "Arsa yatırımında en önemli kriter nedir?", "options": ["Lokasyon", "Fiyat", "Büyüklük", "Şekil"], "correct_answer": "Lokasyon"},
                    {"id": "q2", "text": "İmar planı hangi kurum tarafından hazırlanır?", "options": ["Belediye", "Tapu Müdürlüğü", "Kadastro", "Maliye"], "correct_answer": "Belediye"},
                    {"id": "q3", "text": "Tapu tescili için hangi belge gereklidir?", "options": ["Nüfus cüzdanı", "Satış vaadi sözleşmesi", "Tapu senedi", "İmar belgesi"], "correct_answer": "Tapu senedi"},
                    {"id": "q4", "text": "Parsel numarası neyi ifade eder?", "options": ["Taşınmaz kimliği", "Bina katı", "Kat planı", "Bina yaşı"], "correct_answer": "Taşınmaz kimliği"},
                    {"id": "q5", "text": "E.M.S.A.L. kısaltması ne anlama gelir?", "options": ["Emsal değeri", "Emsalsiz Mülk Satış Alanı", "Emsal Miktarı Standart Arazi Listesi", "Emsalli"], "correct_answer": "Emsal değeri"},
                ]
            }
            await db.course_exams.insert_one(exam)

@app.on_event("startup")
async def startup():
    try:
        init_storage()
        logger.info("Object Storage initialized successfully")
    except Exception as e:
        logger.error(f"Object Storage init failed: {e}")
    # Ensure admin user exists (always update password to stay in sync)
    try:
        await db.users.update_one(
            {"email": "ipatarazi@gmail.com"},
            {"$set": {
                "email": "ipatarazi@gmail.com",
                "full_name": "Admin",
                "password": hash_password("As537273"),
                "role": "admin",
            }},
            upsert=True
        )
        logger.info("Admin user ensured")
    except Exception as e:
        logger.error(f"Admin seed error: {e}")

    # ---- Seed all collections from exported data if empty ----
    SEED_COLLECTIONS = [
        "projects", "shared_facilities", "project_map_layers", "project_media",
        "courses", "seminars", "land_parcels", "land_opportunities",
        "mega_projects", "market_data", "edu_page_settings", "weekly_live",
        "app_users", "yatirim_fonu_basvurulari", "yatirim_fonu_bekleme",
        "seo_settings",
    ]
    for col_name in SEED_COLLECTIONS:
        try:
            col = db[col_name]
            count = await col.count_documents({})
            if count == 0 and col_name in SEED_DATA and SEED_DATA[col_name]:
                docs = SEED_DATA[col_name]
                await col.insert_many(docs)
                logger.info(f"Seeded {col_name}: {len(docs)} documents")
            elif count > 0:
                logger.info(f"{col_name}: {count} mevcut, seed atlandı")
        except Exception as e:
            logger.error(f"Seed error for {col_name}: {e}")

    # Seed exams for user panel demo
    try:
        await seed_exams()
    except Exception as e:
        logger.error(f"Exam seed error: {e}")

# CRITICAL: Add CORS middleware BEFORE including router
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(api_router)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

