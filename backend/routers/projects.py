from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Response, Query
from fastapi.responses import Response as FastAPIResponse
from datetime import datetime, timezone
from typing import Optional
import uuid, io, json

from database import db
from auth_utils import require_admin
from storage import put_object, get_object, MIME_TYPES, APP_NAME

router = APIRouter()

PROJECT_TYPES = ["TOKİ", "Emlak Konut", "Özel Proje", "Kamu Projesi"]
MEDIA_CATEGORIES = ["Altyapı", "Blok Resimleri", "Peyzaj", "Zemin", "Drone", "Master Plan"]
DOC_TYPES = ["Zemin Etüt", "Jeoloji / Jeoteknik", "ÇED", "İhale Belgeleri", "Plan Notları", "Vaziyet Planı", "Diğer"]


def extract_centroid_from_kml(fext: str, data: bytes):
    import re, xml.etree.ElementTree as ET, zipfile as _zf, io as _io2
    try:
        kml_bytes = data
        if fext == "kmz":
            with _zf.ZipFile(_io2.BytesIO(data)) as z2:
                kfiles = [n for n in z2.namelist() if n.lower().endswith('.kml')]
                if not kfiles:
                    return None
                kml_bytes = z2.read(kfiles[0])
        kml_text = kml_bytes.decode('utf-8', errors='ignore')
        coords_raw = ' '.join(
            el.text or '' for el in ET.fromstring(kml_text).iter()
            if el.tag.endswith('coordinates')
        )
        pairs = re.findall(r'([-\d.]+),([-\d.]+)', coords_raw)
        if pairs:
            lons = [float(p[0]) for p in pairs]
            lats = [float(p[1]) for p in pairs]
            return {"lat": round(sum(lats)/len(lats), 6), "lng": round(sum(lons)/len(lons), 6)}
    except Exception:
        pass
    return None


def extract_centroid_from_geojson(data: bytes):
    import json as _json
    all_lons, all_lats = [], []

    def collect(coords):
        if not coords:
            return
        if isinstance(coords[0], (int, float)):
            if len(coords) >= 2:
                all_lons.append(float(coords[0]))
                all_lats.append(float(coords[1]))
        else:
            for item in coords:
                collect(item)

    def process_geom(geom):
        if not geom:
            return
        gtype = geom.get('type', '')
        if gtype == 'GeometryCollection':
            for g in geom.get('geometries', []):
                process_geom(g)
        else:
            collect(geom.get('coordinates', []))

    try:
        geojson = _json.loads(data.decode('utf-8'))
        t = geojson.get('type', '')
        if t == 'FeatureCollection':
            for f in geojson.get('features', []):
                process_geom(f.get('geometry'))
        elif t == 'Feature':
            process_geom(geojson.get('geometry'))
        else:
            process_geom(geojson)
        if all_lats:
            return {"lat": round(sum(all_lats) / len(all_lats), 6), "lng": round(sum(all_lons) / len(all_lons), 6)}
    except Exception:
        pass
    return None


# ============= PROJECTS CRUD =============

@router.post("/admin/projects")
async def create_project(
    admin: dict = Depends(require_admin),
    project_name: str = Form(...), city: str = Form(...), district: str = Form(...),
    neighborhood: str = Form(""), description: str = Form(""), project_type: str = Form("TOKİ"),
    total_housing: int = Form(0), commercial_count: int = Form(0), school_count: int = Form(0),
    mosque_count: int = Form(0), social_facility_count: int = Form(0),
    project_area_sqm: float = Form(0), start_date: str = Form(""), planned_end_date: str = Form(""),
    progress_percentage: int = Form(0), location_lat: float = Form(41.0082), location_lng: float = Form(28.9784),
):
    project = {
        "id": str(uuid.uuid4()), "project_name": project_name, "city": city, "district": district,
        "neighborhood": neighborhood, "description": description, "project_type": project_type,
        "total_housing": total_housing, "commercial_count": commercial_count, "school_count": school_count,
        "mosque_count": mosque_count, "social_facility_count": social_facility_count,
        "project_area_sqm": project_area_sqm, "start_date": start_date,
        "planned_end_date": planned_end_date, "progress_percentage": progress_percentage,
        "location": {"lat": location_lat, "lng": location_lng}, "youtube_videos": [],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.projects.insert_one(project)
    project.pop("_id", None)
    return project


@router.get("/projects")
async def get_projects(city: Optional[str] = None, district: Optional[str] = None, project_name: Optional[str] = None, project_type: Optional[str] = None):
    query = {}
    if city: query["city"] = {"$regex": city, "$options": "i"}
    if district: query["district"] = {"$regex": district, "$options": "i"}
    if project_name: query["project_name"] = {"$regex": project_name, "$options": "i"}
    if project_type: query["project_type"] = project_type
    return await db.projects.find(query, {"_id": 0}).limit(500).to_list(500)


@router.get("/projects/{project_id}")
async def get_project(project_id: str):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project


@router.put("/admin/projects/{project_id}")
async def update_project(
    project_id: str, admin: dict = Depends(require_admin),
    project_name: str = Form(...), city: str = Form(...), district: str = Form(...),
    neighborhood: str = Form(""), description: str = Form(""), project_type: str = Form("TOKİ"),
    total_housing: int = Form(0), commercial_count: int = Form(0), school_count: int = Form(0),
    mosque_count: int = Form(0), social_facility_count: int = Form(0),
    project_area_sqm: float = Form(0), start_date: str = Form(""), planned_end_date: str = Form(""),
    progress_percentage: int = Form(0), location_lat: float = Form(41.0082), location_lng: float = Form(28.9784),
):
    update_data = {
        "project_name": project_name, "city": city, "district": district,
        "neighborhood": neighborhood, "description": description, "project_type": project_type,
        "total_housing": total_housing, "commercial_count": commercial_count,
        "school_count": school_count, "mosque_count": mosque_count,
        "social_facility_count": social_facility_count, "project_area_sqm": project_area_sqm,
        "start_date": start_date, "planned_end_date": planned_end_date,
        "progress_percentage": progress_percentage,
        "location": {"lat": location_lat, "lng": location_lng},
    }
    result = await db.projects.update_one({"id": project_id}, {"$set": update_data})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return await db.projects.find_one({"id": project_id}, {"_id": 0})


@router.delete("/admin/projects/{project_id}")
async def delete_project(project_id: str, admin: dict = Depends(require_admin)):
    result = await db.projects.delete_one({"id": project_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    await db.project_adas.delete_many({"project_id": project_id})
    await db.project_parsels.delete_many({"project_id": project_id})
    await db.project_media.delete_many({"project_id": project_id})
    await db.project_documents.delete_many({"project_id": project_id})
    await db.project_map_layers.delete_many({"project_id": project_id})
    return {"message": "Project deleted"}


# ============= ADA CRUD =============

@router.post("/admin/projects/{project_id}/adas")
async def create_ada(project_id: str, admin: dict = Depends(require_admin), ada_no: str = Form(...), description: str = Form("")):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    existing = await db.project_adas.find_one({"project_id": project_id, "ada_no": ada_no})
    if existing:
        raise HTTPException(status_code=400, detail=f"Ada {ada_no} already exists in this project")
    ada = {
        "id": str(uuid.uuid4()), "project_id": project_id, "ada_no": ada_no,
        "description": description, "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_adas.insert_one(ada)
    ada.pop("_id", None)
    return ada


@router.get("/projects/{project_id}/adas")
async def get_adas(project_id: str):
    return await db.project_adas.find({"project_id": project_id}, {"_id": 0}).sort("ada_no", 1).limit(500).to_list(500)


@router.put("/admin/adas/{ada_id}")
async def update_ada(ada_id: str, admin: dict = Depends(require_admin), ada_no: str = Form(...), description: str = Form("")):
    result = await db.project_adas.update_one({"id": ada_id}, {"$set": {"ada_no": ada_no, "description": description}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ada not found")
    return await db.project_adas.find_one({"id": ada_id}, {"_id": 0})


@router.delete("/admin/adas/{ada_id}")
async def delete_ada(ada_id: str, admin: dict = Depends(require_admin)):
    ada = await db.project_adas.find_one({"id": ada_id})
    if not ada:
        raise HTTPException(status_code=404, detail="Ada not found")
    await db.project_adas.delete_one({"id": ada_id})
    await db.project_parsels.delete_many({"ada_id": ada_id})
    return {"message": "Ada and its parcels deleted"}


# ============= PARSEL CRUD =============

@router.post("/admin/adas/{ada_id}/parsels")
async def create_parsel(ada_id: str, admin: dict = Depends(require_admin), parsel_no: str = Form(...), area_sqm: float = Form(0), note: str = Form("")):
    ada = await db.project_adas.find_one({"id": ada_id})
    if not ada:
        raise HTTPException(status_code=404, detail="Ada not found")
    parsel = {
        "id": str(uuid.uuid4()), "project_id": ada["project_id"], "ada_id": ada_id,
        "parsel_no": parsel_no, "area_sqm": area_sqm if area_sqm > 0 else None,
        "note": note, "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_parsels.insert_one(parsel)
    parsel.pop("_id", None)
    return parsel


@router.get("/projects/{project_id}/parsels")
async def get_parsels(project_id: str):
    return await db.project_parsels.find({"project_id": project_id}, {"_id": 0}).limit(5000).to_list(5000)


@router.put("/admin/parsels/{parsel_id}")
async def update_parsel(parsel_id: str, admin: dict = Depends(require_admin), parsel_no: str = Form(...), area_sqm: float = Form(0), note: str = Form("")):
    result = await db.project_parsels.update_one({"id": parsel_id}, {"$set": {"parsel_no": parsel_no, "area_sqm": area_sqm if area_sqm > 0 else None, "note": note}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Parsel not found")
    return await db.project_parsels.find_one({"id": parsel_id}, {"_id": 0})


@router.delete("/admin/parsels/{parsel_id}")
async def delete_parsel(parsel_id: str, admin: dict = Depends(require_admin)):
    result = await db.project_parsels.delete_one({"id": parsel_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Parsel not found")
    return {"message": "Parsel deleted"}


# ============= EXCEL IMPORT =============

def _parse_excel_rows(content: bytes, ext: str):
    """Parse Excel or CSV bytes and return list of row dicts."""
    if ext == "csv":
        import csv
        reader = csv.DictReader(io.StringIO(content.decode("utf-8-sig")))
        return list(reader)
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val
        rows.append(row_dict)
    return rows


@router.post("/admin/projects/{project_id}/import-excel")
async def import_excel(project_id: str, file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Only xlsx, xls, csv files are supported")
    content = await file.read()
    errors = []
    success_count = 0
    try:
        rows = _parse_excel_rows(content, ext)
        for idx, row in enumerate(rows, start=2):
            ada_no = str(row.get("Ada", "") or "").strip()
            parsel_no = str(row.get("Parsel", "") or "").strip()
            area_str = row.get("Alan_m2", "")
            note = str(row.get("Not", "") or "").strip()
            if not ada_no or not parsel_no:
                errors.append({"row": idx, "error": "Ada veya Parsel boş"})
                continue
            area_sqm = None
            if area_str:
                try:
                    area_sqm = float(area_str)
                except (ValueError, TypeError):
                    errors.append({"row": idx, "error": f"Geçersiz alan değeri: {area_str}"})
                    continue
            ada = await db.project_adas.find_one({"project_id": project_id, "ada_no": ada_no})
            if not ada:
                ada = {
                    "id": str(uuid.uuid4()), "project_id": project_id, "ada_no": ada_no,
                    "description": "", "created_at": datetime.now(timezone.utc).isoformat(),
                }
                await db.project_adas.insert_one(ada)
            parsel = {
                "id": str(uuid.uuid4()), "project_id": project_id, "ada_id": ada["id"],
                "parsel_no": parsel_no, "area_sqm": area_sqm, "note": note,
                "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.project_parsels.insert_one(parsel)
            success_count += 1
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"File processing error: {str(e)}")
    return {"success_count": success_count, "error_count": len(errors), "errors": errors}


@router.get("/admin/excel-template")
async def download_excel_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ada-Parsel"
    ws.append(["Ada", "Parsel", "Alan_m2", "Not"])
    ws.append(["33", "1", "500", "Örnek not"])
    ws.append(["33", "2", "750", ""])
    ws.append(["34", "7", "1200", "Köşe parsel"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FastAPIResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ada_parsel_sablonu.xlsx"}
    )


# ============= PROJECT EXCEL IMPORT =============

@router.get("/admin/project-excel-template")
async def download_project_excel_template():
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projeler"
    headers = ["Proje_Adi", "Il", "Ilce", "Mahalle", "Proje_Tipi", "Aciklama",
               "Konut_Sayisi", "Ticari_Alan", "Okul", "Cami", "Sosyal_Tesis",
               "Proje_Alani_m2", "Baslangic_Tarihi", "Bitis_Tarihi", "Ilerleme_Yuzde", "Enlem", "Boylam"]
    ws.append(headers)
    ws.append(["Örnek Proje", "İstanbul", "Arnavutköy", "Tayakadın", "TOKİ", "Açıklama metni",
               "761", "197", "1", "1", "2", "50000", "2024-01-15", "2026-06-30", "45", "41.0082", "28.9784"])
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(max_len + 2, 12)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FastAPIResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=proje_sablonu.xlsx"}
    )


@router.post("/admin/projects/import-excel")
async def import_projects_excel(file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Sadece xlsx, xls, csv desteklenir")
    content = await file.read()
    errors = []
    success_count = 0
    created_ids = []

    def safe_int(val, default=0):
        try: return int(float(val)) if val else default
        except (ValueError, TypeError): return default

    def safe_float(val, default=0.0):
        try: return float(val) if val else default
        except (ValueError, TypeError): return default

    try:
        rows = _parse_excel_rows(content, ext)
        for idx, row in enumerate(rows, start=2):
            name = str(row.get("Proje_Adi", "") or "").strip()
            city = str(row.get("Il", "") or "").strip()
            if not name or not city:
                errors.append({"row": idx, "error": "Proje_Adi veya Il boş"})
                continue
            project = {
                "id": str(uuid.uuid4()), "project_name": name, "city": city,
                "district": str(row.get("Ilce", "") or "").strip(),
                "neighborhood": str(row.get("Mahalle", "") or "").strip(),
                "description": str(row.get("Aciklama", "") or "").strip(),
                "project_type": str(row.get("Proje_Tipi", "TOKİ") or "TOKİ").strip(),
                "total_housing": safe_int(row.get("Konut_Sayisi")),
                "commercial_count": safe_int(row.get("Ticari_Alan")),
                "school_count": safe_int(row.get("Okul")),
                "mosque_count": safe_int(row.get("Cami")),
                "social_facility_count": safe_int(row.get("Sosyal_Tesis")),
                "project_area_sqm": safe_float(row.get("Proje_Alani_m2")),
                "start_date": str(row.get("Baslangic_Tarihi", "") or "").strip(),
                "planned_end_date": str(row.get("Bitis_Tarihi", "") or "").strip(),
                "progress_percentage": safe_int(row.get("Ilerleme_Yuzde")),
                "location": {"lat": safe_float(row.get("Enlem"), 41.0082), "lng": safe_float(row.get("Boylam"), 28.9784)},
                "youtube_videos": [], "created_at": datetime.now(timezone.utc).isoformat(),
            }
            await db.projects.insert_one(project)
            project.pop("_id", None)
            created_ids.append(project["id"])
            success_count += 1
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Dosya işleme hatası: {str(e)}")
    return {"success_count": success_count, "error_count": len(errors), "errors": errors, "created_ids": created_ids}


# ============= MEDIA MANAGEMENT =============

@router.post("/admin/projects/{project_id}/media")
async def upload_media(project_id: str, file: UploadFile = File(...), category: str = Form(...), admin: dict = Depends(require_admin)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "bin"
    storage_path = f"{APP_NAME}/media/{project_id}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    content_type = file.content_type or MIME_TYPES.get(ext, "application/octet-stream")
    put_object(storage_path, data, content_type)
    media = {
        "id": str(uuid.uuid4()), "project_id": project_id, "category": category,
        "storage_path": storage_path, "original_filename": file.filename,
        "content_type": content_type, "size": len(data),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_media.insert_one(media)
    media.pop("_id", None)
    return media


@router.get("/projects/{project_id}/media")
async def get_project_media(project_id: str, category: Optional[str] = None):
    query = {"project_id": project_id}
    if category:
        query["category"] = category
    return await db.project_media.find(query, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)


@router.delete("/admin/media/{media_id}")
async def delete_media(media_id: str, admin: dict = Depends(require_admin)):
    result = await db.project_media.delete_one({"id": media_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Media not found")
    return {"message": "Media deleted"}


# ============= DOCUMENT MANAGEMENT =============

@router.post("/admin/projects/{project_id}/documents")
async def upload_document(project_id: str, file: UploadFile = File(...), title: str = Form(...), doc_type: str = Form(...), admin: dict = Depends(require_admin)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else "bin"
    storage_path = f"{APP_NAME}/documents/{project_id}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    content_type = file.content_type or MIME_TYPES.get(ext, "application/octet-stream")
    put_object(storage_path, data, content_type)
    doc = {
        "id": str(uuid.uuid4()), "project_id": project_id, "title": title, "doc_type": doc_type,
        "storage_path": storage_path, "original_filename": file.filename,
        "content_type": content_type, "size": len(data),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_documents.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.get("/projects/{project_id}/documents")
async def get_project_documents(project_id: str):
    return await db.project_documents.find({"project_id": project_id}, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)


@router.delete("/admin/documents/{doc_id}")
async def delete_document(doc_id: str, admin: dict = Depends(require_admin)):
    result = await db.project_documents.delete_one({"id": doc_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"message": "Document deleted"}


# ============= VIDEO MANAGEMENT =============

@router.post("/admin/projects/{project_id}/videos")
async def add_video(project_id: str, admin: dict = Depends(require_admin), youtube_url: str = Form(...)):
    result = await db.projects.update_one({"id": project_id}, {"$push": {"youtube_videos": youtube_url}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"message": "Video added"}


@router.delete("/admin/projects/{project_id}/videos")
async def remove_video(project_id: str, admin: dict = Depends(require_admin), youtube_url: str = Form(...)):
    result = await db.projects.update_one({"id": project_id}, {"$pull": {"youtube_videos": youtube_url}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Project not found")
    return {"message": "Video removed"}


# ============= MAP LAYERS =============

@router.post("/admin/projects/{project_id}/map-layers")
async def upload_map_layer(project_id: str, file: UploadFile = File(...), admin: dict = Depends(require_admin)):
    project = await db.projects.find_one({"id": project_id})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    if ext not in ("kml", "kmz", "geojson", "json"):
        raise HTTPException(status_code=400, detail="Only KML, KMZ, GeoJSON files are supported")
    storage_path = f"{APP_NAME}/map-layers/{project_id}/{uuid.uuid4()}.{ext}"
    data = await file.read()
    content_type = file.content_type or MIME_TYPES.get(ext, "application/octet-stream")
    put_object(storage_path, data, content_type)
    layer = {
        "id": str(uuid.uuid4()), "project_id": project_id, "storage_path": storage_path,
        "original_filename": file.filename, "content_type": content_type,
        "file_type": ext.upper(), "size": len(data),
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.project_map_layers.insert_one(layer)
    centroid = None
    try:
        if ext in ("kml", "kmz"):
            centroid = extract_centroid_from_kml(ext, data)
        elif ext in ("geojson", "json"):
            centroid = extract_centroid_from_geojson(data)
    except Exception:
        pass
    if centroid:
        await db.projects.update_one({"id": project_id}, {"$set": {"location": centroid}})
        layer["location_updated"] = centroid
    layer.pop("_id", None)
    return layer


@router.get("/projects/{project_id}/map-layers")
async def get_map_layers(project_id: str):
    return await db.project_map_layers.find({"project_id": project_id}, {"_id": 0}).to_list(100)


@router.delete("/admin/map-layers/{layer_id}")
async def delete_map_layer(layer_id: str, admin: dict = Depends(require_admin)):
    result = await db.project_map_layers.delete_one({"id": layer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Map layer not found")
    return {"message": "Map layer deleted"}


# ============= FILE SERVING =============

@router.get("/files/{path:path}")
async def serve_file(path: str):
    try:
        data, content_type = get_object(path)
        return FastAPIResponse(content=data, media_type=content_type)
    except Exception:
        raise HTTPException(status_code=404, detail="File not found")


# ============= MAP LAYER DATA =============

@router.get("/projects/{project_id}/map-layers/{layer_id}/data")
async def get_map_layer_data(project_id: str, layer_id: str):
    layer = await db.project_map_layers.find_one({"id": layer_id, "project_id": project_id}, {"_id": 0})
    if not layer:
        raise HTTPException(status_code=404, detail="Map layer not found")
    try:
        data, _ = get_object(layer["storage_path"])
        if layer["file_type"] in ("GEOJSON", "JSON"):
            return json.loads(data)
        if layer["file_type"] == "KMZ":
            import zipfile
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                kml_files = [n for n in z.namelist() if n.endswith('.kml')]
                if not kml_files:
                    raise HTTPException(status_code=400, detail="No KML found inside KMZ")
                kml_content = z.read(kml_files[0])
            return FastAPIResponse(content=kml_content, media_type="application/vnd.google-earth.kml+xml")
        return FastAPIResponse(content=data, media_type=layer["content_type"])
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Cannot read layer: {str(e)}")


# ============= BACKWARD COMPAT: TOKI ENDPOINTS =============

@router.get("/toki/projects")
async def get_toki_projects(city: Optional[str] = None, district: Optional[str] = None, project_name: Optional[str] = None):
    query = {}
    if city: query["city"] = {"$regex": city, "$options": "i"}
    if district: query["district"] = {"$regex": district, "$options": "i"}
    if project_name: query["project_name"] = {"$regex": project_name, "$options": "i"}
    return await db.projects.find(query, {"_id": 0}).limit(500).to_list(500)


@router.get("/toki/projects/{project_id}")
async def get_toki_project(project_id: str):
    project = await db.projects.find_one({"id": project_id}, {"_id": 0})
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return project
