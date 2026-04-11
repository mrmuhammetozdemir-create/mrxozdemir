import os
"""
Tests for Project Excel Import Feature
- GET /api/admin/project-excel-template - downloads Excel template with project columns
- POST /api/admin/projects/import-excel - imports projects from Excel file
"""
import pytest
import requests
import os
import io

BASE_URL = os.environ.get('REACT_APP_BACKEND_URL', '').rstrip('/')

@pytest.fixture(scope="module")
def auth_token():
    """Get authentication token"""
    response = requests.post(f"{BASE_URL}/api/auth/login", json={
        "email": "ipatarazi@gmail.com",
        "password": "As537273"
    })
    assert response.status_code == 200, f"Login failed: {response.text}"
    return response.json()["access_token"]

@pytest.fixture(scope="module")
def auth_headers(auth_token):
    """Headers with auth token"""
    return {"Authorization": f"Bearer {auth_token}"}

class TestProjectExcelTemplate:
    """Tests for GET /api/admin/project-excel-template"""
    
    def test_template_download_success(self):
        """Template downloads successfully with correct content type"""
        response = requests.get(f"{BASE_URL}/api/admin/project-excel-template")
        assert response.status_code == 200
        assert "spreadsheetml.sheet" in response.headers.get("Content-Type", "")
        assert "attachment" in response.headers.get("Content-Disposition", "")
        # Verify file is not empty (should be > 1KB)
        assert len(response.content) > 1000
        
    def test_template_filename(self):
        """Template has correct filename in Content-Disposition"""
        response = requests.get(f"{BASE_URL}/api/admin/project-excel-template")
        assert response.status_code == 200
        disposition = response.headers.get("Content-Disposition", "")
        assert "proje_sablonu.xlsx" in disposition
        
    def test_template_is_valid_excel(self):
        """Template is a valid Excel file with expected columns"""
        import openpyxl
        response = requests.get(f"{BASE_URL}/api/admin/project-excel-template")
        assert response.status_code == 200
        
        # Parse as Excel file
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Check headers exist
        headers = [str(cell.value or "") for cell in ws[1]]
        expected_headers = ["Proje_Adi", "Il", "Ilce", "Mahalle", "Proje_Tipi", "Aciklama",
                          "Konut_Sayisi", "Ticari_Alan", "Okul", "Cami", "Sosyal_Tesis",
                          "Proje_Alani_m2", "Baslangic_Tarihi", "Bitis_Tarihi", "Ilerleme_Yuzde",
                          "Enlem", "Boylam"]
        
        for expected in expected_headers:
            assert expected in headers, f"Missing column: {expected}"

class TestProjectExcelImport:
    """Tests for POST /api/admin/projects/import-excel"""
    
    def test_import_requires_auth(self):
        """Import endpoint requires authentication"""
        # Create a simple Excel file
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Proje_Adi", "Il", "Ilce"])
        ws.append(["Test", "Istanbul", "Kadikoy"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        # Try without auth
        response = requests.post(
            f"{BASE_URL}/api/admin/projects/import-excel",
            files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        assert response.status_code in [401, 403]
        
    def test_import_rejects_invalid_file_type(self, auth_headers):
        """Import rejects non-Excel files"""
        response = requests.post(
            f"{BASE_URL}/api/admin/projects/import-excel",
            headers=auth_headers,
            files={"file": ("test.txt", b"not an excel file", "text/plain")}
        )
        assert response.status_code == 400
        assert "xlsx" in response.json().get("detail", "").lower() or "xls" in response.json().get("detail", "").lower()
        
    def test_import_single_valid_project(self, auth_headers):
        """Import single valid project succeeds"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["Proje_Adi", "Il", "Ilce", "Mahalle", "Proje_Tipi", "Aciklama",
                   "Konut_Sayisi", "Ticari_Alan", "Okul", "Cami", "Sosyal_Tesis",
                   "Proje_Alani_m2", "Baslangic_Tarihi", "Bitis_Tarihi", "Ilerleme_Yuzde",
                   "Enlem", "Boylam"]
        ws.append(headers)
        ws.append(["TEST_Single_Import_Project", "İzmir", "Konak", "Alsancak", "TOKİ", 
                   "Single test project", 200, 30, 1, 0, 1, 15000, "2024-05-01", "2025-12-31", 25, 38.4, 27.1])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/admin/projects/import-excel",
            headers=auth_headers,
            files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["success_count"] == 1
        assert data["error_count"] == 0
        assert len(data["created_ids"]) == 1
        
        # Verify project was created
        project_id = data["created_ids"][0]
        get_response = requests.get(f"{BASE_URL}/api/projects/{project_id}")
        assert get_response.status_code == 200
        project = get_response.json()
        assert project["project_name"] == "TEST_Single_Import_Project"
        assert project["city"] == "İzmir"
        assert project["district"] == "Konak"
        assert project["total_housing"] == 200
        assert project["project_type"] == "TOKİ"
        
    def test_import_multiple_projects(self, auth_headers):
        """Import multiple projects at once"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["Proje_Adi", "Il", "Ilce", "Konut_Sayisi"]
        ws.append(headers)
        ws.append(["TEST_Multi_Import_1", "İstanbul", "Beşiktaş", 100])
        ws.append(["TEST_Multi_Import_2", "Ankara", "Etimesgut", 200])
        ws.append(["TEST_Multi_Import_3", "Bursa", "Nilüfer", 300])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/admin/projects/import-excel",
            headers=auth_headers,
            files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["success_count"] == 3
        assert data["error_count"] == 0
        assert len(data["created_ids"]) == 3
        
    def test_import_handles_missing_proje_adi(self, auth_headers):
        """Import reports error for missing Proje_Adi"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Proje_Adi", "Il", "Ilce"])
        ws.append(["", "İstanbul", "Kadıköy"])  # Missing project name
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/admin/projects/import-excel",
            headers=auth_headers,
            files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["success_count"] == 0
        assert data["error_count"] == 1
        assert len(data["errors"]) == 1
        assert data["errors"][0]["row"] == 2
        assert "Proje_Adi" in data["errors"][0]["error"] or "Il" in data["errors"][0]["error"]
        
    def test_import_handles_missing_il(self, auth_headers):
        """Import reports error for missing Il (city)"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Proje_Adi", "Il", "Ilce"])
        ws.append(["TEST_Missing_City", "", "Kadıköy"])  # Missing city
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/admin/projects/import-excel",
            headers=auth_headers,
            files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["success_count"] == 0
        assert data["error_count"] == 1
        
    def test_import_mixed_valid_invalid_rows(self, auth_headers):
        """Import processes valid rows and reports errors for invalid ones"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Proje_Adi", "Il", "Ilce", "Konut_Sayisi"])
        ws.append(["TEST_Mixed_Valid_1", "İstanbul", "Kadıköy", 100])  # Valid
        ws.append(["", "Ankara", "Çankaya", 200])  # Invalid - no name
        ws.append(["TEST_Mixed_Valid_2", "İzmir", "Bornova", 150])  # Valid
        ws.append(["TEST_Missing_City", "", "Nilüfer", 175])  # Invalid - no city
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/admin/projects/import-excel",
            headers=auth_headers,
            files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["success_count"] == 2
        assert data["error_count"] == 2
        assert len(data["created_ids"]) == 2
        assert len(data["errors"]) == 2
        
    def test_import_csv_file(self, auth_headers):
        """Import accepts CSV files"""
        csv_content = """Proje_Adi,Il,Ilce,Konut_Sayisi
TEST_CSV_Import_Project,İstanbul,Üsküdar,250"""
        
        response = requests.post(
            f"{BASE_URL}/api/admin/projects/import-excel",
            headers=auth_headers,
            files={"file": ("test.csv", csv_content.encode("utf-8-sig"), "text/csv")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["success_count"] == 1
        assert data["error_count"] == 0
        
    def test_import_preserves_all_fields(self, auth_headers):
        """Import correctly maps all Excel columns to project fields"""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["Proje_Adi", "Il", "Ilce", "Mahalle", "Proje_Tipi", "Aciklama",
                   "Konut_Sayisi", "Ticari_Alan", "Okul", "Cami", "Sosyal_Tesis",
                   "Proje_Alani_m2", "Baslangic_Tarihi", "Bitis_Tarihi", "Ilerleme_Yuzde",
                   "Enlem", "Boylam"]
        ws.append(headers)
        ws.append(["TEST_Full_Fields_Project", "İstanbul", "Arnavutköy", "Tayakadın", "Emlak Konut",
                   "Detailed description here", 761, 197, 2, 1, 3, 50000,
                   "2024-01-15", "2026-06-30", 45, 41.0082, 28.9784])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        
        response = requests.post(
            f"{BASE_URL}/api/admin/projects/import-excel",
            headers=auth_headers,
            files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        )
        
        assert response.status_code == 200
        data = response.json()
        assert data["success_count"] == 1
        project_id = data["created_ids"][0]
        
        # Fetch and verify all fields
        get_response = requests.get(f"{BASE_URL}/api/projects/{project_id}")
        assert get_response.status_code == 200
        project = get_response.json()
        
        assert project["project_name"] == "TEST_Full_Fields_Project"
        assert project["city"] == "İstanbul"
        assert project["district"] == "Arnavutköy"
        assert project["neighborhood"] == "Tayakadın"
        assert project["project_type"] == "Emlak Konut"
        assert project["description"] == "Detailed description here"
        assert project["total_housing"] == 761
        assert project["commercial_count"] == 197
        assert project["school_count"] == 2
        assert project["mosque_count"] == 1
        assert project["social_facility_count"] == 3
        assert project["project_area_sqm"] == 50000
        assert project["start_date"] == "2024-01-15"
        assert project["planned_end_date"] == "2026-06-30"
        assert project["progress_percentage"] == 45
        assert abs(project["location"]["lat"] - 41.0082) < 0.001
        assert abs(project["location"]["lng"] - 28.9784) < 0.001

@pytest.fixture(scope="module", autouse=True)
def cleanup_test_projects(auth_headers):
    """Cleanup TEST_ prefixed projects after all tests"""
    yield
    # Delete test projects
    response = requests.get(f"{BASE_URL}/api/projects")
    if response.status_code == 200:
        projects = response.json()
        for project in projects:
            if project.get("project_name", "").startswith("TEST_"):
                requests.delete(f"{BASE_URL}/api/admin/projects/{project['id']}", headers=auth_headers)

if __name__ == "__main__":
    pytest.main([__file__, "-v"])
